"""
build_fuel_db.py
Reads the daily stock report Excel files from both fuel companies and
builds the consolidated Fuel.xlsx database used by the dashboard.

Usage:
    python app/build_fuel_db.py

Source files (update paths below if the files move):
  - Pacific Energy:   Daily stock report - Pacific Energy 2104.xlsx
  - TotalEnergies:    Daily stock report - TotalEnergies_04052026 (002).xlsx
"""

import openpyxl
import pandas as pd
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SRC_DIR   = Path(r"C:\Users\halea\NextCloud\Documents\Petroleum\Fuel Crisis\Dashboard")
PACIFIC   = SRC_DIR / "Daily stock report - Pacific Energy 2104.xlsx"
TOTAL_EN  = SRC_DIR / "Daily stock report - TotalEnergies_04052026 (002).xlsx"
OUTPUT    = Path(__file__).resolve().parent.parent / "data" / "Fuel.xlsx"
TODAY     = pd.Timestamp.today().normalize()


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------
def _is_real_date(v):
    return isinstance(v, datetime)

def _num(v):
    try:
        return float(v) if v not in (None, "#VALUE!", "#N/A", "") else None
    except Exception:
        return None


# ---------------------------------------------------------------------------
# 1. Pacific Energy — Tongatapu
#    Layout (0-indexed cols):
#      col 7 = date | 8=Petrol CS | 9=Petrol Offtake | 10=Diesel CS |
#      11=Diesel Offtake | 12=Kerosene CS | 13=Kerosene Offtake
#    Resupply (left block):
#      col 0=label | 1=date | 2=Petrol qty | 3=Diesel qty | 4=Kerosene qty
# ---------------------------------------------------------------------------
def parse_pacific_tongatapu():
    wb = openpyxl.load_workbook(str(PACIFIC), read_only=True, data_only=True)
    ws = wb["TONGATAPU Daily Stock Report"]
    actual_rows, resupply_rows = [], []
    for row in ws.iter_rows(values_only=True):
        # --- Actual ---
        date = row[7] if len(row) > 7 else None
        if _is_real_date(date) and pd.Timestamp(date) <= TODAY:
            for fuel, cs_col, oft_col in [("Petrol", 8, 9), ("Diesel", 10, 11), ("Kerosene", 12, 13)]:
                cs  = _num(row[cs_col])  if len(row) > cs_col  else None
                oft = _num(row[oft_col]) if len(row) > oft_col else None
                if cs is not None:
                    actual_rows.append({
                        "Date": pd.Timestamp(date),
                        "Closing Stock": cs,
                        "Offtake": oft or 0,
                        "Fuel Type": fuel,
                        "Company": "Pacific Energy",
                        "Location": "Tongatapu",
                        "Tonga Power Offtake": 0,
                    })
        # --- Resupply ---
        label = str(row[0]) if row[0] else ""
        if "resupply" in label.lower() and _is_real_date(row[1] if len(row) > 1 else None):
            r_date = pd.Timestamp(row[1])
            for fuel, qty_col in [("Petrol", 2), ("Diesel", 3), ("Kerosene", 4)]:
                qty = _num(row[qty_col]) if len(row) > qty_col else None
                if qty:
                    resupply_rows.append({
                        "Date": r_date, "Quantity": qty,
                        "Fuel Type": fuel,
                        "Company": "Pacific Energy", "Location": "Tongatapu",
                    })
    wb.close()
    return actual_rows, resupply_rows


# ---------------------------------------------------------------------------
# 2. Pacific Energy — Vava'u
#    Layout (0-indexed cols):
#      col 6 = date | 7=Petrol CS | 8=Petrol Offtake |
#      9=Diesel CS | 10=Diesel Offtake
#    Resupply: col 0=label | 1=date | 2=Petrol qty | 3=Diesel qty
# ---------------------------------------------------------------------------
def parse_pacific_vavau():
    wb = openpyxl.load_workbook(str(PACIFIC), read_only=True, data_only=True)
    ws = wb["VAVA'U Daily Stock Report"]
    actual_rows, resupply_rows = [], []
    for row in ws.iter_rows(values_only=True):
        # --- Actual ---
        date = row[6] if len(row) > 6 else None
        if _is_real_date(date) and pd.Timestamp(date) <= TODAY:
            for fuel, cs_col, oft_col in [("Petrol", 7, 8), ("Diesel", 9, 10)]:
                cs  = _num(row[cs_col])  if len(row) > cs_col  else None
                oft = _num(row[oft_col]) if len(row) > oft_col else None
                if cs is not None:
                    actual_rows.append({
                        "Date": pd.Timestamp(date),
                        "Closing Stock": cs,
                        "Offtake": oft or 0,
                        "Fuel Type": fuel,
                        "Company": "Pacific Energy",
                        "Location": "Vava'u",
                        "Tonga Power Offtake": 0,
                    })
        # --- Resupply ---
        label = str(row[0]) if row[0] else ""
        if "resupply" in label.lower() and _is_real_date(row[1] if len(row) > 1 else None):
            r_date = pd.Timestamp(row[1])
            for fuel, qty_col in [("Petrol", 2), ("Diesel", 3)]:
                qty = _num(row[qty_col]) if len(row) > qty_col else None
                if qty:
                    resupply_rows.append({
                        "Date": r_date, "Quantity": qty,
                        "Fuel Type": fuel,
                        "Company": "Pacific Energy", "Location": "Vava'u",
                    })
    wb.close()
    return actual_rows, resupply_rows


# ---------------------------------------------------------------------------
# 3. TotalEnergies — Tongatapu
#    Layout (0-indexed cols):
#      col 6 = date | 7=Petrol CS | 8=Petrol Offtake |
#      9=Diesel CS | 10=Diesel Offtake-TongaPower | 11=Diesel Offtake-Remainder | 12=Diesel Offtake-Total
#    Resupply: col 0=label | 1=date | 2=Petrol qty | 3=Diesel qty
# ---------------------------------------------------------------------------
def parse_totalenergies():
    wb = openpyxl.load_workbook(str(TOTAL_EN), read_only=True, data_only=True)
    ws = wb["Daily Stock"]
    actual_rows, resupply_rows = [], []
    for row in ws.iter_rows(values_only=True):
        # --- Actual ---
        date = row[6] if len(row) > 6 else None
        if _is_real_date(date) and pd.Timestamp(date) <= TODAY:
            petrol_cs  = _num(row[7])  if len(row) > 7  else None
            petrol_oft = _num(row[8])  if len(row) > 8  else None
            diesel_cs  = _num(row[9])  if len(row) > 9  else None
            tp_oft     = _num(row[10]) if len(row) > 10 else None  # Tonga Power offtake
            diesel_oft = _num(row[12]) if len(row) > 12 else None  # Total diesel offtake

            if petrol_cs is not None:
                actual_rows.append({
                    "Date": pd.Timestamp(date),
                    "Closing Stock": petrol_cs,
                    "Offtake": petrol_oft or 0,
                    "Fuel Type": "Petrol",
                    "Company": "TotalEnergies",
                    "Location": "Tongatapu",
                    "Tonga Power Offtake": 0,
                })
            if diesel_cs is not None:
                actual_rows.append({
                    "Date": pd.Timestamp(date),
                    "Closing Stock": diesel_cs,
                    "Offtake": diesel_oft or 0,
                    "Fuel Type": "Diesel",
                    "Company": "TotalEnergies",
                    "Location": "Tongatapu",
                    "Tonga Power Offtake": tp_oft or 0,
                })

        # --- Resupply ---
        label = str(row[0]) if row[0] else ""
        if "resupply" in label.lower() and _is_real_date(row[1] if len(row) > 1 else None):
            r_date = pd.Timestamp(row[1])
            for fuel, qty_col in [("Petrol", 2), ("Diesel", 3)]:
                qty = _num(row[qty_col]) if len(row) > qty_col else None
                if qty:
                    resupply_rows.append({
                        "Date": r_date, "Quantity": qty,
                        "Fuel Type": fuel,
                        "Company": "TotalEnergies", "Location": "Tongatapu",
                    })
    wb.close()
    return actual_rows, resupply_rows


# ---------------------------------------------------------------------------
# 4. Build & write
# ---------------------------------------------------------------------------
def main():
    all_actual, all_resupply = [], []

    for fn in [parse_pacific_tongatapu, parse_pacific_vavau, parse_totalenergies]:
        a, r = fn()
        all_actual.extend(a)
        all_resupply.extend(r)

    actual   = pd.DataFrame(all_actual)
    resupply = pd.DataFrame(all_resupply) if all_resupply else pd.DataFrame(
        columns=["Date", "Quantity", "Fuel Type", "Company", "Location"])
    terminal = pd.DataFrame(columns=["Date", "Quantity", "Fuel Type", "Company", "Location"])

    # Preserve existing FuelPrice_Long and Tariff_Long
    fuel_price = pd.DataFrame(columns=["Date", "Fuel Type", "Price"])
    tariff     = pd.DataFrame(columns=["Date", "Value"])
    if OUTPUT.exists():
        try:
            xls = pd.ExcelFile(str(OUTPUT))
            if "FuelPrice_Long" in xls.sheet_names:
                fuel_price = pd.read_excel(str(OUTPUT), sheet_name="FuelPrice_Long")
            if "Tariff_Long" in xls.sheet_names:
                tariff = pd.read_excel(str(OUTPUT), sheet_name="Tariff_Long")
        except Exception as e:
            print(f"Warning: could not preserve price/tariff sheets: {e}")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(str(OUTPUT), engine="openpyxl") as writer:
        actual.to_excel(writer,   sheet_name="Actual",         index=False)
        resupply.to_excel(writer, sheet_name="Resupply",       index=False)
        terminal.to_excel(writer, sheet_name="Terminal",       index=False)
        fuel_price.to_excel(writer, sheet_name="FuelPrice_Long", index=False)
        tariff.to_excel(writer,   sheet_name="Tariff_Long",    index=False)

    print(f"Written to: {OUTPUT}")
    print(f"  Actual rows   : {len(actual)}")
    if not actual.empty:
        print(f"  Date range    : {actual['Date'].min().date()} → {actual['Date'].max().date()}")
        print(f"  Companies     : {sorted(actual['Company'].unique())}")
        print(f"  Locations     : {sorted(actual['Location'].unique())}")
        print(f"  Fuel types    : {sorted(actual['Fuel Type'].unique())}")
    print(f"  Resupply rows : {len(resupply)}")
    print(f"  FuelPrice rows: {len(fuel_price)}")
    print(f"  Tariff rows   : {len(tariff)}")


if __name__ == "__main__":
    main()
