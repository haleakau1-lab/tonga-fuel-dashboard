import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from pathlib import Path
import numpy as np
import base64
import io
import hashlib
import json
from datetime import datetime, timedelta

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

st.set_page_config(page_title="Tonga National Fuel Supply & Consumption Dashboard", layout="wide")

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_FILE = BASE_DIR / "Oil_Data_Consolidated.xlsx"
PRICE_FILE = BASE_DIR / "Transformed_for_Analysis.xlsx"
LOGO_FILE = BASE_DIR / "tonga-energy-logo.jpg"
BACKGROUND_FILE = BASE_DIR / "background.jpg"
FUEL_COMPANY_DIR = BASE_DIR / "Fuel Company"
FORECAST_ENTRY_FILE = BASE_DIR / "Forecast_Entries.csv"
CHART_COLORS = ["#38BDF8", "#F97316", "#34D399", "#FBBF24", "#A78BFA", "#FB7185"]

# Authentication credentials (can be changed or moved to environment variables)
CREDENTIALS_FILE = BASE_DIR / "credentials.json"

# Default credentials - will be used if credentials.json doesn't exist
DEFAULT_CREDENTIALS = {
    "admin": "admin123",  # username: password
    "operator": "operator123"
}


def hash_password(password):
    """Hash password for storage."""
    return hashlib.sha256(password.encode()).hexdigest()


def load_credentials():
    """Load credentials from file or return defaults."""
    if CREDENTIALS_FILE.exists():
        try:
            with open(CREDENTIALS_FILE, 'r') as f:
                loaded = json.load(f)
            normalized = {}
            for username, password in loaded.items():
                password_text = str(password)
                is_hash = len(password_text) == 64 and all(ch in "0123456789abcdef" for ch in password_text.lower())
                normalized[str(username)] = password_text if is_hash else hash_password(password_text)
            return normalized
        except Exception:
            pass

    return {username: hash_password(password) for username, password in DEFAULT_CREDENTIALS.items()}


def authenticate_user(username, password):
    """Validate a username and password against the configured credential store."""
    credentials = load_credentials()
    return credentials.get(str(username).strip()) == hash_password(password)


def init_session_state():
    """Initialize authentication-related session state."""
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if "username" not in st.session_state:
        st.session_state.username = None
    if "auth_time" not in st.session_state:
        st.session_state.auth_time = None


def check_session_timeout(timeout_hours=8):
    """Return True when the login session is valid and unexpired."""
    if not st.session_state.get("authenticated"):
        return False

    auth_time = st.session_state.get("auth_time")
    if not auth_time:
        return False

    if datetime.now() - auth_time > timedelta(hours=timeout_hours):
        st.session_state.authenticated = False
        st.session_state.username = None
        st.session_state.auth_time = None
        st.warning("Your session has expired. Please sign in again.")
        return False

    return True


def show_login_form():
    """Render the login form for the protected data-entry page."""
    left, center, right = st.columns([1.25, 1.6, 1.25])
    with center:
        with st.container(border=True):
            st.markdown("### Data Entry Login")
            st.caption("Sign in to access protected data entry forms.")
            with st.form("login_form", border=False):
                username = st.text_input("Username")
                password = st.text_input("Password", type="password")
                submitted = st.form_submit_button("Login", use_container_width=True)

    if submitted:
        if authenticate_user(username, password):
            st.session_state.authenticated = True
            st.session_state.username = username.strip()
            st.session_state.auth_time = datetime.now()
            st.success("Login successful")
            st.rerun()
        else:
            st.error("Invalid username or password")


@st.cache_data
def load_data(file_obj, file_mtime=None):
    xls = pd.ExcelFile(file_obj)
    actual = pd.read_excel(file_obj, sheet_name="Actual")
    resupply = pd.read_excel(file_obj, sheet_name="Resupply")
    terminal = pd.read_excel(file_obj, sheet_name="Terminal")

    # Normalize dtypes for reliable filtering and charting.
    if "Date" in actual.columns:
        actual["Date"] = pd.to_datetime(actual["Date"], errors="coerce")
    if "Date" in resupply.columns:
        resupply["Date"] = pd.to_datetime(resupply["Date"], errors="coerce")

    for col in ["Closing Stock", "Offtake", "Tonga Power Offtake"]:
        if col in actual.columns:
            actual[col] = pd.to_numeric(actual[col], errors="coerce")

    if "Quantity" in resupply.columns:
        resupply["Quantity"] = pd.to_numeric(resupply["Quantity"], errors="coerce")

    if "Quantity" in terminal.columns:
        terminal["Quantity"] = pd.to_numeric(terminal["Quantity"], errors="coerce")

    for frame in (actual, resupply, terminal):
        if "Company" in frame.columns:
            frame["Company"] = frame["Company"].apply(_normalize_company_name)

    return xls.sheet_names, actual, resupply, terminal


@st.cache_data
def load_price_data(file_path, mtime=None):
    fp = pd.read_excel(file_path, sheet_name="FuelPrice_Long")
    tr = pd.read_excel(file_path, sheet_name="Tariff_Long")
    fp["Date"] = pd.to_datetime(fp["Date"], errors="coerce")
    fp["Price"] = pd.to_numeric(fp["Price"], errors="coerce")
    tr["Value"] = pd.to_numeric(tr["Value"], errors="coerce")

    # Normalize tariff date for files that store Month/Year instead of Date.
    if "Date" not in tr.columns:
        if {"Year", "Month"}.issubset(tr.columns):
            month_numbers = pd.to_datetime(tr["Month"], format="%b", errors="coerce").dt.month
            if month_numbers.isna().all():
                month_numbers = pd.to_datetime(tr["Month"], format="%B", errors="coerce").dt.month
            year_start = tr["Year"].astype(str).str.extract(r"(\d{4})", expand=False)
            tr["Date"] = pd.to_datetime(
                year_start + "-" + month_numbers.fillna(1).astype(int).astype(str).str.zfill(2) + "-01",
                errors="coerce",
            )
        elif "Month" in tr.columns:
            tr["Date"] = pd.to_datetime(tr["Month"], errors="coerce")
    else:
        tr["Date"] = pd.to_datetime(tr["Date"], errors="coerce")

    return fp, tr


def _company_from_filename(file_name):
    name = file_name.lower()
    if "totalenergies" in name:
        return "TotalEnergies"
    if "total energies" in name:
        return "TotalEnergies"
    if "pacific energy" in name:
        return "Pacific Energy"
    if "co-1" in name:
        return "TotalEnergies"
    if "co-2" in name:
        return "Pacific Energy"
    return Path(file_name).stem


def _normalize_company_name(raw_company):
    text = str(raw_company).strip()
    normalized = text.lower().replace("-", " ")
    normalized = " ".join(normalized.split())

    if normalized in {"company 1", "co 1", "co1"}:
        return "TotalEnergies"
    if normalized in {"company 2", "co 2", "co2"}:
        return "Pacific Energy"
    if "totalenergies" in normalized or "total energies" in normalized:
        return "TotalEnergies"
    if "pacific energy" in normalized:
        return "Pacific Energy"
    return text


def _company_allowed(company_name):
    return str(company_name).strip() in {"TotalEnergies", "Pacific Energy"}


def _location_from_sheet(sheet_name):
    s = str(sheet_name).upper()
    if "VAVA" in s:
        return "Vava'u"
    if "TONGATAPU" in s:
        return "Tongatapu"
    return str(sheet_name)


def _normalize_fuel_name(raw_fuel):
    text = str(raw_fuel).strip()
    text = text.replace("(litres)", "").replace("(Litres)", "").strip()
    if not text:
        return "Unknown"
    return text


@st.cache_data
def load_forecast_from_company_files(folder_path):
    from openpyxl import load_workbook

    records = []
    folder = Path(folder_path)
    if not folder.exists():
        return pd.DataFrame(columns=["Date", "Company", "Location", "Fuel Type", "Forecast Delivery", "Forecast Closing", "Source"])

    files = sorted([p for p in folder.glob("*.xlsx") if not p.name.startswith("~$")])
    for file_path in files:
        company = _company_from_filename(file_path.name)
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception:
            continue

        for ws in wb.worksheets:
            location = _location_from_sheet(ws.title)

            update_row = None
            fuels = []
            for r in range(1, min(ws.max_row, 40) + 1):
                row_first_cell = str(ws.cell(r, 1).value or "").strip().lower()
                if row_first_cell == "update as required":
                    update_row = r
                    for c in range(3, 10):
                        fuel_val = ws.cell(r, c).value
                        if fuel_val:
                            fuels.append(_normalize_fuel_name(fuel_val))
                    break

            header_row = None
            for r in range(1, min(ws.max_row, 40) + 1):
                row_values = [str(ws.cell(r, c).value or "").lower() for c in range(1, min(ws.max_column, 80) + 1)]
                if any("forecast delivery" in value for value in row_values):
                    header_row = r
                    break

            if header_row is None:
                continue

            forecast_pairs = []
            for c in range(1, min(ws.max_column, 120)):
                left = str(ws.cell(header_row, c).value or "").strip().lower()
                right = str(ws.cell(header_row, c + 1).value or "").strip().lower()
                if "forecast delivery" in left and "closing stock" in right:
                    forecast_pairs.append((c, c + 1))

            if not forecast_pairs:
                continue

            date_col = None
            if update_row is not None:
                for c in range(1, min(ws.max_column, 30) + 1):
                    if str(ws.cell(update_row, c).value or "").strip().lower() == "date":
                        date_col = c
                        break
            if date_col is None:
                date_col = 7

            start_row = (update_row + 1) if update_row is not None else (header_row + 1)
            end_row = min(ws.max_row, start_row + 450)

            for r in range(start_row, end_row + 1):
                raw_date = ws.cell(r, date_col).value
                date_value = pd.to_datetime(raw_date, errors="coerce")
                if pd.isna(date_value):
                    continue

                for i, (delivery_col, closing_col) in enumerate(forecast_pairs):
                    fuel_name = fuels[i] if i < len(fuels) else f"Fuel {i + 1}"
                    delivery = pd.to_numeric(ws.cell(r, delivery_col).value, errors="coerce")
                    closing = pd.to_numeric(ws.cell(r, closing_col).value, errors="coerce")
                    if pd.isna(delivery) and pd.isna(closing):
                        continue
                    records.append(
                        {
                            "Date": date_value,
                            "Company": company,
                            "Location": location,
                            "Fuel Type": fuel_name,
                            "Forecast Delivery": delivery,
                            "Forecast Closing": closing,
                            "Source": f"{file_path.name} | {ws.title}",
                        }
                    )

    if not records:
        return pd.DataFrame(columns=["Date", "Company", "Location", "Fuel Type", "Forecast Delivery", "Forecast Closing", "Source"])

    forecast = pd.DataFrame(records)
    forecast["Date"] = pd.to_datetime(forecast["Date"], errors="coerce")
    if "Company" in forecast.columns:
        forecast["Company"] = forecast["Company"].apply(_normalize_company_name)
    forecast["Forecast Delivery"] = pd.to_numeric(forecast["Forecast Delivery"], errors="coerce")
    forecast["Forecast Closing"] = pd.to_numeric(forecast["Forecast Closing"], errors="coerce")
    forecast = forecast.dropna(subset=["Date"]).sort_values("Date")
    return forecast


@st.cache_data
def load_manual_forecast_entries(file_path):
    fp = Path(file_path)
    if not fp.exists():
        return pd.DataFrame(columns=["Date", "Company", "Location", "Fuel Type", "Forecast Delivery", "Forecast Closing", "Source"])

    manual = pd.read_csv(fp)
    if manual.empty:
        return pd.DataFrame(columns=["Date", "Company", "Location", "Fuel Type", "Forecast Delivery", "Forecast Closing", "Source"])

    manual["Date"] = pd.to_datetime(manual.get("Date"), errors="coerce")
    if "Company" in manual.columns:
        manual["Company"] = manual["Company"].apply(_normalize_company_name)
    manual["Forecast Delivery"] = pd.to_numeric(manual.get("Forecast Delivery"), errors="coerce")
    manual["Forecast Closing"] = pd.to_numeric(manual.get("Forecast Closing"), errors="coerce")
    if "Source" not in manual.columns:
        manual["Source"] = "Manual Entry"
    return manual[["Date", "Company", "Location", "Fuel Type", "Forecast Delivery", "Forecast Closing", "Source"]]


def calculate_kpis(actual_df, resupply_df):
    """Calculate KPI metrics."""
    # Total Stock: sum of all latest closing stock values
    latest_actual = actual_df.dropna(subset=["Date", "Closing Stock"])
    if not latest_actual.empty:
        latest_actual = latest_actual.sort_values("Date")
        latest_actual = latest_actual.drop_duplicates(subset=["Company", "Location", "Fuel Type"], keep="last")
        total_stock = latest_actual["Closing Stock"].sum()
    else:
        total_stock = 0

    # Total Offtake excludes Tonga Power so the two KPI cards do not double count.
    offtake_data = actual_df.copy()
    offtake_data["Offtake"] = pd.to_numeric(offtake_data.get("Offtake"), errors="coerce").fillna(0)
    offtake_data["Tonga Power Offtake"] = pd.to_numeric(
        offtake_data.get("Tonga Power Offtake"), errors="coerce"
    ).fillna(0)
    total_offtake = (offtake_data["Offtake"] - offtake_data["Tonga Power Offtake"]).clip(lower=0).sum()
    total_consumption = offtake_data["Offtake"].sum()

    # Upcoming Supply: sum of all resupply quantities
    upcoming_data = resupply_df.dropna(subset=["Quantity"])
    upcoming_supply = upcoming_data["Quantity"].sum()

    return total_stock, total_offtake, upcoming_supply, total_consumption


def calculate_stock_by_fuel(actual_df):
    """Calculate latest stock by fuel type."""
    latest_actual = actual_df.dropna(subset=["Date", "Closing Stock"])
    if not latest_actual.empty:
        latest_actual = latest_actual.sort_values("Date")
        latest_actual = latest_actual.drop_duplicates(subset=["Company", "Location", "Fuel Type"], keep="last")
        stock_by_fuel = latest_actual.groupby("Fuel Type")["Closing Stock"].sum()
        return stock_by_fuel
    return pd.Series()


def calculate_days_of_cover(total_stock, daily_offtake):
    """Calculate days of cover."""
    if daily_offtake > 0:
        return total_stock / daily_offtake
    return 0


def calculate_cover_status(days_of_cover):
    """Classify stock risk status based on days of cover."""
    if days_of_cover >= 45:
        return "Safe"
    if days_of_cover >= 30:
        return "Watch"
    return "Critical"


def to_csv(df):
    """Convert DataFrame to CSV bytes."""
    return df.to_csv(index=False).encode("utf-8")


def checkbox_slicer(container, title, options, key_prefix):
    """Render a checkbox-style slicer and return selected options."""
    container.markdown(f'<div class="filter-title-chip">{title}</div>', unsafe_allow_html=True)

    clear_pending_key = f"{key_prefix}_clear_pending"

    for idx, _ in enumerate(options):
        option_key = f"{key_prefix}_{idx}"
        if option_key not in st.session_state:
            st.session_state[option_key] = True

    # Apply clear action before rendering checkbox widgets to avoid
    # modifying widget-bound keys after instantiation.
    if st.session_state.get(clear_pending_key, False):
        for idx, _ in enumerate(options):
            st.session_state[f"{key_prefix}_{idx}"] = False
        st.session_state[clear_pending_key] = False

    selected = []

    for idx, option in enumerate(options):
        checked = container.checkbox(
            str(option),
            key=f"{key_prefix}_{idx}",
        )
        if checked:
            selected.append(option)

    # Bottom-right action controls for each filter group.
    if container.button("Clear", key=f"{key_prefix}_clear"):
        st.session_state[clear_pending_key] = True
        st.rerun()

    return selected


def checkbox_slicer_horizontal(container, title, options, key_prefix):
    """Render checkbox options in a single horizontal row and return selected items."""
    container.markdown(f'<div class="filter-title-chip">{title}</div>', unsafe_allow_html=True)

    clear_pending_key = f"{key_prefix}_clear_pending"

    for idx, _ in enumerate(options):
        option_key = f"{key_prefix}_{idx}"
        if option_key not in st.session_state:
            st.session_state[option_key] = True

    if st.session_state.get(clear_pending_key, False):
        for idx, _ in enumerate(options):
            st.session_state[f"{key_prefix}_{idx}"] = False
        st.session_state[clear_pending_key] = False

    selected = []
    if options:
        row_cols = container.columns([1] * len(options) + [0.7], gap="small")
        for idx, option in enumerate(options):
            with row_cols[idx]:
                checked = st.checkbox(str(option), key=f"{key_prefix}_{idx}")
            if checked:
                selected.append(option)

        with row_cols[-1]:
            if st.button("Clear", key=f"{key_prefix}_clear"):
                st.session_state[clear_pending_key] = True
                st.rerun()
    else:
        if container.button("Clear", key=f"{key_prefix}_clear"):
            st.session_state[clear_pending_key] = True
            st.rerun()

    return selected


def format_compact(value):
    """Format numbers using compact K/M suffixes for KPI cards."""
    if pd.isna(value):
        return "0"
    abs_value = abs(float(value))
    if abs_value >= 1_000_000:
        return f"{value / 1_000_000:.1f}M"
    if abs_value >= 1_000:
        return f"{value / 1_000:.0f}K"
    return f"{value:.0f}"


def fuel_icon(fuel_name):
    """Return a compact icon per fuel label for KPI readability."""
    fuel_text = str(fuel_name).strip().lower()
    if "diesel" in fuel_text:
        return "🛢️"
    if "petrol" in fuel_text or "gasoline" in fuel_text:
        return "⛽"
    if "kerosene" in fuel_text or "jet" in fuel_text:
        return "✈️"
    return "🔹"


def render_kpi_tile(container, icon, label, value, accent):
    """Render a compact KPI tile with a colored accent for better scanability."""
    container.markdown(
        f"""
        <div class="kpi-tile" style="border-left: 3px solid {accent};">
            <div class="kpi-label">{icon} {label}</div>
            <div class="kpi-value">{value}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_kpi_category(container, title, items, class_name=""):
    """Render a compact category card with tightly controlled KPI spacing."""
    items_html = "".join(
        [
            f'<div class="kpi-mini" style="border-left: 3px solid {accent};">'
            f'<div class="kpi-mini-line">'
            f'<div class="kpi-mini-label">{icon} {label}</div>'
            f'<div class="kpi-mini-value">{value}</div>'
            f'</div>'
            f'</div>'
            for icon, label, value, accent in items
        ]
    )
    container.markdown(
        (
            f'<div class="kpi-category-card {class_name}">'
            f'<div class="kpi-category-title">{title}</div>'
            f'<div class="kpi-mini-row">{items_html}</div>'
            f'</div>'
        ),
        unsafe_allow_html=True,
    )


def render_kpi_group(container, title, items):
    """Render one grouped KPI block with a single border around title and cards."""
    cards_html = "".join(
        [
            (
                f'<div class="kpi-group-card" style="border-left: 3px solid {accent};">'
                f'<div class="kpi-group-card-label">{icon} {label}</div>'
                f'<div class="kpi-group-card-value">{value}</div>'
                f'</div>'
            )
            for icon, label, value, accent in items
        ]
    )
    container.markdown(
        (
            '<div class="kpi-group-frame">'
            f'<div class="kpi-group-title">{title}</div>'
            f'<div class="kpi-group-row">{cards_html}</div>'
            '</div>'
        ),
        unsafe_allow_html=True,
    )


def render_chart_title(container, title):
    """Render a highlighted chart title bar inside chart containers."""
    container.markdown(f'<div class="chart-title-bar">{title}</div>', unsafe_allow_html=True)

def set_app_background(image_path):
    """Set dashboard background image from a local file."""
    if not image_path.exists():
        return

    encoded = base64.b64encode(image_path.read_bytes()).decode("utf-8")
    st.markdown(
        f"""
        <style>
        .stApp {{
            background-image: linear-gradient(rgba(8, 13, 20, 0.44), rgba(8, 13, 20, 0.44)), url("data:image/jpeg;base64,{encoded}");
            background-size: cover;
            background-position: center;
            background-attachment: fixed;
        }}

        /* Dark theme: even stronger overlay for max contrast */
        @media (prefers-color-scheme: dark) {{
            .stApp {{
                background-image: linear-gradient(rgba(5, 9, 14, 0.55), rgba(5, 9, 14, 0.55)), url("data:image/jpeg;base64,{encoded}");
            }}
        }}

        [data-testid="stHeader"] {{
            background: rgba(0, 0, 0, 0);
        }}
        [data-testid="stToolbar"] {{
            right: 0.5rem;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def apply_chart_theme(fig, height=400, hovermode=None, x_title=None, y_title=None, date_x=False):
    """Apply a consistent high-contrast chart style and readable formatting."""
    fig.update_layout(
        height=height,
        colorway=CHART_COLORS,
        paper_bgcolor="rgba(0, 0, 0, 0)",
        plot_bgcolor="rgba(12, 18, 26, 0.86)",
        font=dict(color="#E6EDF5", size=14),
        title=dict(x=0.02, xanchor="left", font=dict(size=18, color="#F8FBFF")),
        margin=dict(l=16, r=16, t=62, b=68),
        bargap=0.18,
        legend=dict(
            orientation="h",
            yanchor="top",
            y=-0.16,
            xanchor="left",
            x=0,
            font=dict(color="#E6EDF5", size=13),
        ),
    )
    if hovermode:
        fig.update_layout(hovermode=hovermode)
    fig.update_xaxes(
        showgrid=True,
        gridcolor="rgba(173, 191, 210, 0.22)",
        zeroline=False,
        linecolor="rgba(173, 191, 210, 0.35)",
        tickfont=dict(color="#E6EDF5", size=13),
        title_font=dict(color="#E6EDF5", size=14),
        title_text=x_title,
        tickformat="%d %b" if date_x else None,
    )
    fig.update_yaxes(
        showgrid=True,
        gridcolor="rgba(173, 191, 210, 0.22)",
        zeroline=False,
        linecolor="rgba(173, 191, 210, 0.35)",
        tickfont=dict(color="#E6EDF5", size=13),
        title_font=dict(color="#E6EDF5", size=14),
        title_text=y_title,
        tickformat=",.0f",
        separatethousands=True,
    )
    fig.update_traces(hoverlabel=dict(font=dict(color="#0B1220")))
    return fig


def _to_normalized_day(value):
    ts = pd.to_datetime(value, errors="coerce")
    if pd.isna(ts):
        return None
    return ts.normalize()


def _sheet_find_entry_row(
    ws,
    entry_date,
    company,
    location,
    fuel_type,
    date_col=1,
    company_col=2,
    location_col=3,
    fuel_col=4,
    start_row=2,
):
    """Find a row by Date+Company+Location+Fuel Type and return its row index."""
    target_day = _to_normalized_day(entry_date)
    if target_day is None:
        return None

    target_company = _normalize_company_name(company)
    target_location = str(location).strip().lower()
    target_fuel = str(fuel_type).strip().lower()

    for row_idx in range(start_row, ws.max_row + 1):
        existing_day = _to_normalized_day(ws.cell(row_idx, date_col).value)
        if existing_day is None:
            continue

        existing_company = _normalize_company_name(ws.cell(row_idx, company_col).value)
        existing_location = str(ws.cell(row_idx, location_col).value or "").strip().lower()
        existing_fuel = str(ws.cell(row_idx, fuel_col).value or "").strip().lower()

        if (
            existing_day == target_day
            and existing_company == target_company
            and existing_location == target_location
            and existing_fuel == target_fuel
        ):
            return row_idx

    return None


def save_actual_data(file_path, date, company, location, fuel_type, closing_stock, offtake, tonga_power_offtake=0, allow_update=False):
    """Add a new row to the Actual sheet in the Excel workbook."""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path)
        ws = wb["Actual"]

        existing_row = _sheet_find_entry_row(ws, date, company, location, fuel_type)
        if existing_row is not None:
            if not allow_update:
                return (
                    False,
                    f"Duplicate blocked: {company} already has an Actual entry for {pd.to_datetime(date).strftime('%Y-%m-%d')} at {location} ({fuel_type}). Enable update to edit it.",
                )

            ws[f"E{existing_row}"] = closing_stock
            ws[f"F{existing_row}"] = offtake
            ws[f"G{existing_row}"] = tonga_power_offtake
            wb.save(file_path)
            return True, "Existing stock entry updated successfully!"
        
        # Find the next empty row
        next_row = ws.max_row + 1
        
        # Add data
        ws[f"A{next_row}"] = date
        ws[f"B{next_row}"] = company
        ws[f"C{next_row}"] = location
        ws[f"D{next_row}"] = fuel_type
        ws[f"E{next_row}"] = closing_stock
        ws[f"F{next_row}"] = offtake
        ws[f"G{next_row}"] = tonga_power_offtake
        
        wb.save(file_path)
        return True, "Stock data added successfully!"
    except Exception as e:
        return False, f"Error saving data: {str(e)}"


def save_resupply_data(file_path, date, company, location, fuel_type, quantity, allow_update=False):
    """Add a new row to the Resupply sheet in the Excel workbook."""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path)
        ws = wb["Resupply"]

        existing_row = _sheet_find_entry_row(ws, date, company, location, fuel_type)
        if existing_row is not None:
            if not allow_update:
                return (
                    False,
                    f"Duplicate blocked: {company} already has a Resupply entry for {pd.to_datetime(date).strftime('%Y-%m-%d')} at {location} ({fuel_type}). Enable update to edit it.",
                )

            ws[f"E{existing_row}"] = quantity
            wb.save(file_path)
            return True, "Existing resupply entry updated successfully!"
        
        # Find the next empty row
        next_row = ws.max_row + 1
        
        # Add data
        ws[f"A{next_row}"] = date
        ws[f"B{next_row}"] = company
        ws[f"C{next_row}"] = location
        ws[f"D{next_row}"] = fuel_type
        ws[f"E{next_row}"] = quantity
        
        wb.save(file_path)
        return True, "Resupply data added successfully!"
    except Exception as e:
        return False, f"Error saving data: {str(e)}"


def save_fuel_price(file_path, date, fuel_type, price_type, price):
    """Add a new row to the FuelPrice_Long sheet in the price workbook."""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path)
        ws = wb["FuelPrice_Long"]
        
        # Find the next empty row
        next_row = ws.max_row + 1
        
        # Add data
        ws[f"A{next_row}"] = date
        ws[f"B{next_row}"] = fuel_type
        ws[f"C{next_row}"] = price_type
        ws[f"D{next_row}"] = price
        
        wb.save(file_path)
        return True, "Fuel price added successfully!"
    except Exception as e:
        return False, f"Error saving price data: {str(e)}"


def save_tariff(file_path, date, component, period, value):
    """Add a new row to the Tariff_Long sheet in the price workbook."""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path)
        ws = wb["Tariff_Long"]
        
        # Find the next empty row
        next_row = ws.max_row + 1
        
        # Add data
        ws[f"A{next_row}"] = date
        ws[f"B{next_row}"] = component
        ws[f"C{next_row}"] = period
        ws[f"D{next_row}"] = value
        
        wb.save(file_path)
        return True, "Tariff data added successfully!"
    except Exception as e:
        return False, f"Error saving tariff data: {str(e)}"


def save_forecast_entry(file_path, date, company, location, fuel_type, forecast_delivery, forecast_closing, allow_update=False):
    """Append a forecast record to a CSV file."""
    try:
        record = pd.DataFrame(
            [
                {
                    "Date": pd.to_datetime(date).strftime("%Y-%m-%d"),
                    "Company": _normalize_company_name(company),
                    "Location": location,
                    "Fuel Type": fuel_type,
                    "Forecast Delivery": forecast_delivery,
                    "Forecast Closing": forecast_closing,
                    "Source": "Manual Entry",
                }
            ]
        )

        fp = Path(file_path)
        if fp.exists():
            existing = pd.read_csv(fp)
            if not existing.empty and {"Date", "Company", "Location", "Fuel Type"}.issubset(existing.columns):
                existing_days = pd.to_datetime(existing["Date"], errors="coerce").dt.normalize()
                existing_companies = existing["Company"].apply(_normalize_company_name)
                existing_locations = existing["Location"].astype(str).str.strip().str.lower()
                existing_fuels = existing["Fuel Type"].astype(str).str.strip().str.lower()
                target_day = _to_normalized_day(date)
                target_company = _normalize_company_name(company)
                target_location = str(location).strip().lower()
                target_fuel = str(fuel_type).strip().lower()

                duplicate_mask = (
                    (existing_days == target_day)
                    & (existing_companies == target_company)
                    & (existing_locations == target_location)
                    & (existing_fuels == target_fuel)
                )
                if target_day is not None and duplicate_mask.any():
                    if not allow_update:
                        return (
                            False,
                            f"Duplicate blocked: {target_company} already has a Forecast entry for {target_day.strftime('%Y-%m-%d')} at {location} ({fuel_type}). Enable update to edit it.",
                        )

                    update_idx = existing.index[duplicate_mask][0]
                    existing.loc[update_idx, "Forecast Delivery"] = forecast_delivery
                    existing.loc[update_idx, "Forecast Closing"] = forecast_closing
                    existing.loc[update_idx, "Source"] = "Manual Entry"
                    existing.to_csv(fp, index=False)
                    return True, "Existing forecast entry updated successfully!"
            updated = pd.concat([existing, record], ignore_index=True)
        else:
            updated = record

        updated.to_csv(fp, index=False)
        return True, "Forecast entry added successfully!"
    except Exception as e:
        return False, f"Error saving forecast data: {str(e)}"


set_app_background(BACKGROUND_FILE)

logo_html = ""
footer_logo_html = ""
if LOGO_FILE.exists():
    logo_b64 = base64.b64encode(LOGO_FILE.read_bytes()).decode("utf-8")
    logo_html = (
        f'<img src="data:image/jpeg;base64,{logo_b64}" '
        'style="height:78px; width:auto; object-fit:contain; border-radius:8px; margin-right:0.75rem; border:1px solid rgba(173, 191, 210, 0.45);" '
        'alt="Organization Logo" />'
    )
    footer_logo_html = (
        f'<img src="data:image/jpeg;base64,{logo_b64}" '
        'style="height:30px; width:auto; object-fit:contain; border-radius:4px;" '
        'alt="DOE Logo" />'
    )

st.markdown(
    f"""
    <div style="
        padding: 0.14rem 0.62rem;
        margin: 0.01rem 0 0.24rem 0;
        border: 1.6px solid rgba(125, 211, 252, 0.72);
        border-radius: 12px;
        background: linear-gradient(180deg, rgba(30, 64, 175, 0.12), rgba(14, 22, 32, 0.66));
        box-shadow: 0 0 0 1px rgba(186, 230, 253, 0.2), 0 0 10px rgba(56, 189, 248, 0.16);
        backdrop-filter: blur(2px);
    ">
        <div style="display:flex; align-items:center; gap:0.25rem;">
            {logo_html}
            <div>
                <h1 style="margin: 0; font-size: 2.2rem; line-height: 1.2; font-weight: 700; letter-spacing: 0.2px;">Tonga National Fuel Supply and Consumption Dashboard</h1>
                <p style="margin: 0.18rem 0 0 0; opacity: 0.88; font-size: 1.08rem; line-height: 1.3;">
                    Staged Energy and Fuel Supply Plan Monitoring
                </p>
            </div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <style>
    .block-container {
        padding-top: 0.5rem;
        padding-bottom: 0.55rem;
        margin-left: auto;
        margin-right: auto;
    }

    div[data-testid="stHorizontalBlock"] {
        gap: 0.5rem;
    }

    .stApp {
        color: #F2F6FA;
        font-size: 16px;
    }

    .stMarkdown p,
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] div {
        font-size: 0.98rem;
    }

    [data-testid="stMarkdownContainer"] h2,
    [data-testid="stMarkdownContainer"] h3 {
        font-size: 1.45rem !important;
        line-height: 1.25;
        font-weight: 700;
        margin-top: 0.08rem;
        margin-bottom: 0.35rem;
        border: 1.8px solid rgba(125, 211, 252, 0.72);
        border-radius: 14px;
        background: linear-gradient(180deg, rgba(30, 64, 175, 0.14), rgba(14, 22, 32, 0.66));
        box-shadow: 0 0 0 1px rgba(186, 230, 253, 0.2), 0 0 10px rgba(56, 189, 248, 0.16);
        padding: 0.26rem 0.62rem;
        color: #F4FAFF;
    }

    [data-testid="stMarkdownContainer"] h3 {
        font-size: 1.18rem !important;
        line-height: 1.3;
    }

    div[data-testid="stMetric"] {
        padding: 0.1rem 0.2rem;
        color: #F2F6FA;
    }

    div[data-testid="stMetricLabel"],
    div[data-testid="stMetricValue"] {
        color: #F2F6FA;
    }

    div[data-testid="stMetricLabel"] {
        font-size: 0.98rem;
        line-height: 1.2;
    }

    div[data-testid="stMetricValue"] {
        font-size: 1.34rem;
        line-height: 1.15;
        font-weight: 500 !important;
    }

    div[data-testid="stCaptionContainer"] {
        font-size: 1.02rem;
        letter-spacing: 0.2px;
    }

    .kpi-tile {
        background: rgba(255, 255, 255, 0.04);
        border: 1px solid rgba(173, 191, 210, 0.18);
        border-radius: 8px;
        padding: 0.45rem 0.5rem;
        min-height: 66px;
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 0.45rem;
        margin-bottom: 0.14rem;
    }

    .kpi-group-title {
        border: none;
        border-radius: 8px;
        background: linear-gradient(90deg, rgba(30, 64, 175, 0.42), rgba(14, 116, 144, 0.36));
        padding: 0.24rem 0.42rem;
        font-size: 0.9rem;
        font-weight: 650;
        line-height: 1.2;
        margin-bottom: 0.35rem;
        text-align: center;
        color: #F8FCFF;
        letter-spacing: 0.2px;
        box-shadow: inset 0 0 0 1px rgba(125, 211, 252, 0.22);
    }

    .kpi-group-frame {
        border: 1.6px solid rgba(125, 211, 252, 0.72);
        border-radius: 14px;
        background: linear-gradient(180deg, rgba(14, 116, 144, 0.1), rgba(14, 22, 32, 0.68));
        box-shadow: 0 0 0 1px rgba(186, 230, 253, 0.2), 0 0 10px rgba(56, 189, 248, 0.16);
        padding: 0.46rem 0.5rem 0.62rem 0.5rem;
    }

    .kpi-group-row {
        display: flex;
        gap: 0.35rem;
        align-items: stretch;
        flex-wrap: wrap;
    }

    .kpi-group-card {
        flex: 1 1 180px;
        background: rgba(255, 255, 255, 0.04);
        border: 1px solid rgba(173, 191, 210, 0.18);
        border-radius: 8px;
        padding: 0.45rem 0.5rem;
        min-height: 66px;
        min-width: 0;
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 0.45rem;
    }

    .kpi-group-card-label {
        font-size: 0.9rem;
        line-height: 1.12;
        opacity: 0.96;
        font-weight: 460 !important;
        white-space: normal;
        overflow: hidden;
        text-overflow: ellipsis;
    }

    .kpi-group-card-value {
        font-size: 1.28rem;
        line-height: 1.05;
        font-weight: 520 !important;
        color: #F8FBFF;
        margin-left: auto;
        white-space: nowrap;
    }

    .chart-title-bar {
        border: 1px solid rgba(125, 211, 252, 0.58);
        border-radius: 10px;
        background: linear-gradient(90deg, rgba(30, 64, 175, 0.22), rgba(14, 116, 144, 0.18));
        color: #F4FAFF;
        font-size: 1rem;
        font-weight: 650;
        line-height: 1.2;
        padding: 0.28rem 0.62rem;
        text-align: center;
        margin: 0.02rem 0 0.38rem 0;
        box-shadow: inset 0 0 0 1px rgba(186, 230, 253, 0.1);
    }

    .st-key-filter_company_group,
    .st-key-filter_location_group,
    .st-key-filter_fuel_group,
    .st-key-filter_month_group,
    .st-key-terminal_location_filter_box,
    .st-key-terminal_type_filter_box,
    .st-key-filter_company_group div[data-testid="stVerticalBlockBorderWrapper"],
    .st-key-filter_location_group div[data-testid="stVerticalBlockBorderWrapper"],
    .st-key-filter_fuel_group div[data-testid="stVerticalBlockBorderWrapper"],
    .st-key-filter_month_group div[data-testid="stVerticalBlockBorderWrapper"],
    .st-key-terminal_location_filter_box div[data-testid="stVerticalBlockBorderWrapper"],
    .st-key-terminal_type_filter_box div[data-testid="stVerticalBlockBorderWrapper"] {
        border: 1.8px solid rgba(125, 211, 252, 0.72) !important;
        border-radius: 14px !important;
        background: linear-gradient(180deg, rgba(30, 64, 175, 0.12), rgba(14, 22, 32, 0.66)) !important;
        box-shadow: 0 0 0 1px rgba(186, 230, 253, 0.2), 0 0 10px rgba(56, 189, 248, 0.16);
        padding: 0.3rem 0.45rem 0.35rem 0.45rem !important;
        margin-bottom: 0.35rem;
    }

    .st-key-chart_stock,
    .st-key-chart_offtake,
    .st-key-chart_location,
    .st-key-chart_resupply,
    .st-key-chart_terminal,
    .st-key-chart_fuel_price,
    .st-key-chart_tariff,
    .st-key-chart_dependency,
    .st-key-chart_stock div[data-testid="stVerticalBlockBorderWrapper"],
    .st-key-chart_offtake div[data-testid="stVerticalBlockBorderWrapper"],
    .st-key-chart_location div[data-testid="stVerticalBlockBorderWrapper"],
    .st-key-chart_resupply div[data-testid="stVerticalBlockBorderWrapper"],
    .st-key-chart_terminal div[data-testid="stVerticalBlockBorderWrapper"],
    .st-key-chart_fuel_price div[data-testid="stVerticalBlockBorderWrapper"],
    .st-key-chart_tariff div[data-testid="stVerticalBlockBorderWrapper"],
    .st-key-chart_dependency div[data-testid="stVerticalBlockBorderWrapper"] {
        border: 1.8px solid rgba(125, 211, 252, 0.72) !important;
        border-radius: 14px !important;
        background: linear-gradient(180deg, rgba(30, 64, 175, 0.12), rgba(14, 22, 32, 0.66)) !important;
        box-shadow: 0 0 0 1px rgba(186, 230, 253, 0.2), 0 0 10px rgba(56, 189, 248, 0.16);
    }

    .filter-title-chip {
        border: 1px solid rgba(125, 211, 252, 0.62);
        border-radius: 8px;
        background: linear-gradient(90deg, rgba(30, 64, 175, 0.2), rgba(14, 116, 144, 0.16));
        color: #F4FAFF;
        font-size: 0.9rem;
        font-weight: 650;
        line-height: 1.2;
        padding: 0.2rem 0.42rem;
        margin: 0.06rem 0 0.3rem 0;
        box-shadow: inset 0 0 0 1px rgba(186, 230, 253, 0.08);
        text-align: center;
    }

    .st-key-filter_company_group div[data-testid="stButton"],
    .st-key-filter_location_group div[data-testid="stButton"],
    .st-key-filter_fuel_group div[data-testid="stButton"],
    .st-key-filter_month_group div[data-testid="stButton"] {
        display: flex;
        justify-content: flex-end;
        margin-top: 0.2rem;
    }

    .st-key-filter_company_group div[data-testid="stButton"] > button,
    .st-key-filter_location_group div[data-testid="stButton"] > button,
    .st-key-filter_fuel_group div[data-testid="stButton"] > button,
    .st-key-filter_month_group div[data-testid="stButton"] > button {
        font-size: 0.76rem;
        line-height: 1.1;
        min-height: 1.68rem;
        min-width: 64px;
        white-space: nowrap;
        padding: 0.08rem 0.42rem;
        border-radius: 7px;
        border: 1px solid rgba(125, 211, 252, 0.78);
        background: linear-gradient(180deg, rgba(30, 64, 175, 0.32), rgba(14, 22, 32, 0.74));
        box-shadow: 0 0 0 1px rgba(186, 230, 253, 0.18), 0 0 8px rgba(56, 189, 248, 0.2);
    }

    .st-key-filter_company_group div[data-testid="stButton"] > button:hover,
    .st-key-filter_location_group div[data-testid="stButton"] > button:hover,
    .st-key-filter_fuel_group div[data-testid="stButton"] > button:hover,
    .st-key-filter_month_group div[data-testid="stButton"] > button:hover {
        border-color: rgba(125, 211, 252, 0.95);
        box-shadow: 0 0 0 1px rgba(186, 230, 253, 0.26), 0 0 12px rgba(56, 189, 248, 0.28);
    }

    /* Highlight the single group-defining border (subtitle + cards together). */
    .st-key-kpi_supply_group,
    .st-key-kpi_offtake_group,
    .st-key-kpi_fuel_group,
    .st-key-kpi_coverage_group {
        border: 2px solid rgba(125, 211, 252, 0.7) !important;
        border-radius: 14px !important;
        box-shadow: 0 0 0 1px rgba(186, 230, 253, 0.26), 0 0 16px rgba(56, 189, 248, 0.22);
        background: linear-gradient(180deg, rgba(14, 116, 144, 0.16), rgba(14, 22, 32, 0.68)) !important;
        padding: 0.34rem 0.34rem 0.56rem 0.34rem !important;
        overflow: hidden;
    }

    .st-key-kpi_supply_group div[data-testid="stVerticalBlockBorderWrapper"],
    .st-key-kpi_offtake_group div[data-testid="stVerticalBlockBorderWrapper"],
    .st-key-kpi_fuel_group div[data-testid="stVerticalBlockBorderWrapper"],
    .st-key-kpi_coverage_group div[data-testid="stVerticalBlockBorderWrapper"] {
        border: none !important;
        box-shadow: none !important;
        background: transparent !important;
        padding: 0.48rem 0.48rem 0.98rem 0.48rem !important;
    }

    .kpi-label {
        font-size: 0.9rem;
        line-height: 1.12;
        opacity: 0.96;
        font-weight: 460 !important;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }

    .kpi-value {
        font-size: 1.28rem;
        line-height: 1.05;
        font-weight: 520 !important;
        color: #F8FBFF;
        margin-left: auto;
        white-space: nowrap;
    }

    .kpi-category-card {
        border: 1px solid rgba(173, 191, 210, 0.38);
        border-radius: 12px;
        background: rgba(14, 22, 32, 0.78);
        backdrop-filter: blur(3px);
        padding: 0.16rem 0.28rem 0.2rem 0.28rem;
    }

    .kpi-category-card.compact-wide {
        max-width: 860px;
        margin-left: auto;
        margin-right: auto;
    }

    .kpi-category-card.compact-wide .kpi-mini-row {
        justify-content: center;
    }

    .kpi-category-card.compact-wide .kpi-mini {
        flex: 0 1 220px;
    }

    .kpi-category-title {
        font-size: 0.98rem;
        font-weight: 500 !important;
        margin: 0 0 0.08rem 0;
        line-height: 1.05;
    }

    .kpi-mini-row {
        display: flex;
        gap: 0.24rem;
        align-items: stretch;
    }

    .kpi-mini {
        flex: 1;
        background: rgba(255, 255, 255, 0.04);
        border-radius: 8px;
        padding: 0.62rem 0.38rem;
        min-height: 100px;
        display: flex;
        align-items: center;
    }

    .kpi-mini-line {
        width: 100%;
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 0.5rem;
    }

    .kpi-mini-label {
        font-size: 0.9rem;
        line-height: 1.1;
        opacity: 0.96;
        font-weight: 460 !important;
        white-space: nowrap;
    }

    .kpi-mini-value {
        margin-top: 0;
        font-size: 1.12rem;
        line-height: 1.05;
        font-weight: 500 !important;
        color: #F8FBFF;
        white-space: nowrap;
    }

    /* Strong override for custom KPI blocks */
    .kpi-category-card .kpi-mini .kpi-mini-value,
    .kpi-category-card .kpi-mini .kpi-mini-label,
    .kpi-category-card .kpi-category-title {
        font-family: "Segoe UI", "Helvetica Neue", Arial, sans-serif;
    }

    /* Compact only the top KPI category boxes */
    .st-key-kpi_supply,
    .st-key-kpi_offtake,
    .st-key-kpi_coverage {
        padding-top: 0;
        padding-bottom: 0;
    }

    .st-key-kpi_supply div[data-testid="stVerticalBlockBorderWrapper"],
    .st-key-kpi_offtake div[data-testid="stVerticalBlockBorderWrapper"],
    .st-key-kpi_coverage div[data-testid="stVerticalBlockBorderWrapper"] {
        padding-top: 0.08rem;
        padding-bottom: 0.2rem;
    }

    .st-key-kpi_supply div[data-testid="stMetric"],
    .st-key-kpi_offtake div[data-testid="stMetric"],
    .st-key-kpi_coverage div[data-testid="stMetric"] {
        padding: 0 0.06rem;
    }

    .st-key-kpi_supply div[data-testid="stMetricValue"],
    .st-key-kpi_offtake div[data-testid="stMetricValue"],
    .st-key-kpi_coverage div[data-testid="stMetricValue"] {
        font-size: 1.12rem;
    }

    .st-key-kpi_supply div[data-testid="stMetricLabel"],
    .st-key-kpi_offtake div[data-testid="stMetricLabel"],
    .st-key-kpi_coverage div[data-testid="stMetricLabel"] {
        font-size: 1rem;
    }

    .st-key-kpi_supply div[data-testid="stCaptionContainer"],
    .st-key-kpi_offtake div[data-testid="stCaptionContainer"],
    .st-key-kpi_coverage div[data-testid="stCaptionContainer"] {
        margin-bottom: 0;
        margin-top: 0;
        line-height: 1.05;
    }

    .st-key-kpi_supply .kpi-tile,
    .st-key-kpi_offtake .kpi-tile,
    .st-key-kpi_coverage .kpi-tile {
        margin-top: -0.05rem;
    }

    div[data-testid="stVerticalBlockBorderWrapper"] {
        background: rgba(14, 22, 32, 0.78);
        border: 1px solid rgba(173, 191, 210, 0.38);
        border-radius: 12px;
        backdrop-filter: blur(3px);
        padding-top: 0.2rem;
        padding-bottom: 0.2rem;
    }

    div[data-testid="stCaptionContainer"] {
        margin-bottom: 0.1rem;
    }

    div[data-testid="stTabs"] {
        margin-top: 0;
    }

    hr {
        margin-top: 0.3rem !important;
        margin-bottom: 0.3rem !important;
    }

    [data-testid="stSidebar"] {
        background: rgba(9, 15, 23, 0.88);
        border-right: 1px solid rgba(173, 191, 210, 0.32);
    }

    [data-testid="stSidebar"] * {
        color: #F2F6FA !important;
    }

    @media (max-width: 900px) {
        .stApp {
            font-size: 15px;
        }
        [data-testid="stMarkdownContainer"] h2 {
            font-size: 1.28rem !important;
        }
        div[data-testid="stMetricValue"] {
            font-size: 1.2rem;
        }
        .kpi-group-frame {
            padding: 0.38rem 0.4rem 0.48rem 0.4rem;
        }
        .kpi-group-row {
            gap: 0.3rem;
        }
        .kpi-group-card {
            flex: 1 1 100%;
            min-height: 58px;
            padding: 0.38rem 0.42rem;
        }
        .kpi-group-card-label {
            font-size: 0.84rem;
            line-height: 1.15;
        }
        .kpi-group-card-value {
            font-size: 1.08rem;
        }
        .kpi-group-title {
            margin-bottom: 0.28rem;
            font-size: 0.86rem;
        }
        .st-key-filter_company_group div[data-testid="stButton"],
        .st-key-filter_location_group div[data-testid="stButton"],
        .st-key-filter_fuel_group div[data-testid="stButton"],
        .st-key-filter_month_group div[data-testid="stButton"] {
            justify-content: stretch;
        }
        .st-key-filter_company_group div[data-testid="stButton"] > button,
        .st-key-filter_location_group div[data-testid="stButton"] > button,
        .st-key-filter_fuel_group div[data-testid="stButton"] > button,
        .st-key-filter_month_group div[data-testid="stButton"] > button {
            width: 100%;
            min-width: 0;
        }
    }

    </style>
    """,
    unsafe_allow_html=True,
)

# Initialize authentication session state
init_session_state()

file_to_use = DEFAULT_FILE

if not DEFAULT_FILE.exists():
    st.error(f"Default file not found at {DEFAULT_FILE}")
    st.stop()

try:
    file_mtime = file_to_use.stat().st_mtime
    sheets, actual_df, resupply_df, terminal_df = load_data(file_to_use, file_mtime=file_mtime)
except Exception as exc:
    st.error(f"Failed to load workbook: {exc}")
    st.stop()

last_sync = pd.to_datetime(file_mtime, unit="s").strftime("%d %b %Y %H:%M")

price_df, tariff_df = None, None
if PRICE_FILE.exists():
    try:
        price_df, tariff_df = load_price_data(str(PRICE_FILE), mtime=PRICE_FILE.stat().st_mtime)
    except Exception:
        pass

forecast_df_auto = load_forecast_from_company_files(str(FUEL_COMPANY_DIR))
forecast_df_manual = load_manual_forecast_entries(str(FORECAST_ENTRY_FILE))
forecast_df = pd.concat([forecast_df_auto, forecast_df_manual], ignore_index=True)
if not forecast_df.empty:
    forecast_df["Date"] = pd.to_datetime(forecast_df["Date"], errors="coerce")
    forecast_df["Forecast Delivery"] = pd.to_numeric(forecast_df["Forecast Delivery"], errors="coerce")
    forecast_df["Forecast Closing"] = pd.to_numeric(forecast_df["Forecast Closing"], errors="coerce")
    forecast_df = forecast_df.dropna(subset=["Date"]).sort_values("Date")

# Sidebar filters
st.sidebar.title("🎛️ Filters")

# Show user info and logout button if authenticated
if st.session_state.authenticated:
    logout_col1, logout_col2 = st.sidebar.columns([2, 1])
    with logout_col1:
        st.caption(f"👤 {st.session_state.username}")
    with logout_col2:
        if st.button("🔓 Logout", key="sidebar_logout", use_container_width=True):
            st.session_state.authenticated = False
            st.session_state.username = None
            st.session_state.auth_time = None
            st.rerun()
    st.sidebar.divider()

_SECTIONS = ["📊 Fuel Supply", "📦 Terminal Data", "💰 Prices & Tariffs", "✏️ Data Entry"]
active_section = st.sidebar.radio("Section", _SECTIONS, key="active_section", label_visibility="collapsed")
st.sidebar.divider()

actual_companies = set(actual_df["Company"].dropna().astype(str).unique()) if "Company" in actual_df.columns else set()
forecast_companies = set(forecast_df["Company"].dropna().astype(str).unique()) if not forecast_df.empty else set()
companies = [company for company in sorted(actual_companies | forecast_companies) if _company_allowed(company)]

actual_locations = set(actual_df["Location"].dropna().astype(str).unique()) if "Location" in actual_df.columns else set()
forecast_locations = set(forecast_df["Location"].dropna().astype(str).unique()) if not forecast_df.empty else set()
locations = sorted(actual_locations | forecast_locations)

actual_fuels = set(actual_df["Fuel Type"].dropna().astype(str).unique()) if "Fuel Type" in actual_df.columns else set()
forecast_fuels = set(forecast_df["Fuel Type"].dropna().astype(str).unique()) if not forecast_df.empty else set()
fuels = sorted(actual_fuels | forecast_fuels)

month_values = []
if "Date" in actual_df.columns:
    month_values.extend(actual_df["Date"].dropna().dt.to_period("M").astype(str).unique().tolist())
if not forecast_df.empty:
    month_values.extend(forecast_df["Date"].dropna().dt.to_period("M").astype(str).unique().tolist())
months = sorted(set(month_values))

if active_section != "💰 Prices & Tariffs":
    company_filter_box = st.sidebar.container(border=True, key="filter_company_group")
    location_filter_box = st.sidebar.container(border=True, key="filter_location_group")
    fuel_filter_box = st.sidebar.container(border=True, key="filter_fuel_group")
    month_filter_box = st.sidebar.container(border=True, key="filter_month_group")
    company_sel = checkbox_slicer(company_filter_box, "Company", companies, "company")
    location_sel = checkbox_slicer(location_filter_box, "Location", locations, "location")
    fuel_sel = checkbox_slicer(fuel_filter_box, "Fuel Type", fuels, "fuel")
    month_sel = checkbox_slicer(month_filter_box, "Month", months, "month")
else:
    company_sel = companies
    location_sel = locations
    fuel_sel = fuels
    month_sel = months
    # Price & Tariff sidebar filters
    _price_opts = sorted(price_df["Price_Type"].dropna().unique().tolist()) if price_df is not None else []
    _fuel_opts = sorted(price_df["Fuel"].dropna().unique().tolist()) if price_df is not None else []
    _comp_opts = ["Fuel Component", "Non Fuel component", "Total Tariff"]
    price_type_filter_box = st.sidebar.container(border=True, key="filter_price_type_group")
    price_fuel_filter_box = st.sidebar.container(border=True, key="filter_price_fuel_group")
    tariff_comp_filter_box = st.sidebar.container(border=True, key="filter_tariff_comp_group")
    price_type_sel = checkbox_slicer(price_type_filter_box, "Price Type", _price_opts, "fp_price_type")
    fuel_sel_fp = checkbox_slicer(price_fuel_filter_box, "Fuel", _fuel_opts, "fp_fuel")
    comp_sel = checkbox_slicer(tariff_comp_filter_box, "Tariff Component", _comp_opts, "tariff_comp")

actual_df_for_filter = actual_df.copy()
if "Date" in actual_df_for_filter.columns:
    actual_df_for_filter["Month"] = actual_df_for_filter["Date"].dt.to_period("M").astype(str)

resupply_df_for_filter = resupply_df.copy()
if "Date" in resupply_df_for_filter.columns:
    resupply_df_for_filter["Month"] = resupply_df_for_filter["Date"].dt.to_period("M").astype(str)

# Apply filters
filtered_actual = actual_df_for_filter[
    actual_df_for_filter["Company"].isin(company_sel)
    & actual_df_for_filter["Location"].isin(location_sel)
    & actual_df_for_filter["Fuel Type"].isin(fuel_sel)
    & actual_df_for_filter["Month"].isin(month_sel)
].copy()

filtered_resupply = resupply_df_for_filter[
    resupply_df_for_filter["Company"].isin(company_sel)
    & resupply_df_for_filter["Location"].isin(location_sel)
    & resupply_df_for_filter["Fuel Type"].isin(fuel_sel)
    & resupply_df_for_filter["Month"].isin(month_sel)
].copy()

forecast_df_for_filter = forecast_df.copy()
if not forecast_df_for_filter.empty:
    forecast_df_for_filter["Month"] = forecast_df_for_filter["Date"].dt.to_period("M").astype(str)
    filtered_forecast = forecast_df_for_filter[
        forecast_df_for_filter["Company"].isin(company_sel)
        & forecast_df_for_filter["Location"].isin(location_sel)
        & forecast_df_for_filter["Fuel Type"].isin(fuel_sel)
        & forecast_df_for_filter["Month"].isin(month_sel)
    ].copy()
else:
    filtered_forecast = forecast_df_for_filter

# Display KPIs
st.subheader("Key Performance Indicators")
total_stock, non_power_offtake, upcoming_supply, total_consumption = calculate_kpis(filtered_actual, filtered_resupply)
tonga_power_offtake = (
    filtered_actual["Tonga Power Offtake"].dropna().sum()
    if "Tonga Power Offtake" in filtered_actual.columns
    else 0
)
avg_daily_offtake = total_consumption / len(filtered_actual["Date"].unique()) if len(filtered_actual["Date"].unique()) > 0 else 0
days_of_cover = calculate_days_of_cover(total_stock, avg_daily_offtake)
cover_status = calculate_cover_status(days_of_cover)

status_icon = "🟢" if cover_status == "Safe" else "🟡" if cover_status == "Watch" else "🔴"
status_accent = "#22C55E" if cover_status == "Safe" else "#F59E0B" if cover_status == "Watch" else "#EF4444"

supply_items = [
    ("🔶", "Total Fuel", f"{format_compact(total_stock)}", "#60A5FA"),
    ("📈", "Upcoming Supply", f"{format_compact(upcoming_supply)}", "#34D399"),
]
offtake_items = [
    ("📊", "Total Offtake", f"{format_compact(total_consumption)}", "#7DD3FC"),
    ("⚡", "Tonga Power", f"{format_compact(tonga_power_offtake)}", "#F59E0B"),
    ("📉", "Non-Power", f"{format_compact(non_power_offtake)}", "#A78BFA"),
]
coverage_items = [
    ("⏱️", "Days", f"{days_of_cover:.2f}", "#38BDF8"),
    (status_icon, "Status", cover_status, status_accent),
]

stock_by_fuel = calculate_stock_by_fuel(filtered_actual)

filter_text = {
    "Company": ", ".join([str(x) for x in company_sel]) if company_sel else "All",
    "Location": ", ".join([str(x) for x in location_sel]) if location_sel else "All",
    "Fuel Type": ", ".join([str(x) for x in fuel_sel]) if fuel_sel else "All",
    "Month": ", ".join([str(x) for x in month_sel]) if month_sel else "All",
}

stock_rows = []
if not stock_by_fuel.empty:
    for fuel_name, stock_value in stock_by_fuel.items():
        stock_rows.append((str(fuel_name), float(stock_value)))

latest_prices = pd.DataFrame(columns=["Price_Type", "Fuel", "Price"])
if price_df is not None and not price_df.empty:
    latest_prices = (
        price_df.dropna(subset=["Date", "Price"])
        .sort_values("Date")
        .drop_duplicates(subset=["Price_Type", "Fuel"], keep="last")
    )

latest_tariff_year = ""
latest_tariff_rows = []
if tariff_df is not None and not tariff_df.empty:
    tariff_main = tariff_df[tariff_df["Component"].isin(["Fuel Component", "Non Fuel component", "Total Tariff"])].copy()
    if not tariff_main.empty:
        latest_tariff_year = sorted(tariff_main["Year"].dropna().astype(str).unique().tolist())[-1]
        tariff_avg = (
            tariff_main[tariff_main["Year"].astype(str) == latest_tariff_year]
            .groupby("Component", as_index=False)["Value"]
            .mean()
        )
        latest_tariff_rows = [(str(r.Component), float(r.Value)) for r in tariff_avg.itertuples(index=False)]


def build_summary_pdf_bytes():
    if not REPORTLAB_AVAILABLE:
        return None

    pdf_buffer = io.BytesIO()
    page_w, page_h = landscape(A4)
    pdf = canvas.Canvas(pdf_buffer, pagesize=(page_w, page_h))

    y = page_h - 30
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(24, y, "Tonga Fuel Dashboard - Weekly Summary")
    y -= 18
    pdf.setFont("Helvetica", 9)
    pdf.drawString(24, y, f"Printed: {pd.Timestamp.now().strftime('%d %b %Y %H:%M')}")
    y -= 12
    pdf.drawString(24, y, f"Data Source: {file_to_use.name} | Last Sync: {last_sync}")
    y -= 12
    pdf.drawString(
        24,
        y,
        (
            f"Filters - Company: {filter_text['Company']} | Location: {filter_text['Location']} | "
            f"Fuel: {filter_text['Fuel Type']} | Month: {filter_text['Month']}"
        ),
    )

    y -= 20
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(24, y, "Key Metrics")
    y -= 14
    pdf.setFont("Helvetica", 9)
    pdf.drawString(28, y, f"Total Fuel: {total_stock:,.0f} L")
    pdf.drawString(220, y, f"Upcoming Supply: {upcoming_supply:,.0f} L")
    pdf.drawString(430, y, f"Total Offtake: {total_consumption:,.0f} L")
    pdf.drawString(620, y, f"Days of Cover: {days_of_cover:.2f} ({cover_status})")

    y -= 22
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(24, y, "Stock by Fuel Type")
    y -= 12
    pdf.setFont("Helvetica", 9)
    for fuel_name, stock_value in stock_rows[:8]:
        pdf.drawString(28, y, f"{fuel_name}")
        pdf.drawRightString(200, y, f"{stock_value:,.0f} L")
        y -= 11
    if not stock_rows:
        pdf.drawString(28, y, "No stock data")
        y -= 11

    y -= 8
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(24, y, "Latest Fuel Prices (Retail/Wholesale)")
    y -= 12
    pdf.setFont("Helvetica", 9)
    if not latest_prices.empty:
        for row in latest_prices.itertuples(index=False):
            pdf.drawString(28, y, f"{row.Price_Type} - {row.Fuel}")
            pdf.drawRightString(260, y, f"{float(row.Price):.2f}")
            y -= 11
            if y < 110:
                break
    else:
        pdf.drawString(28, y, "No fuel price data")
        y -= 11

    y -= 8
    pdf.setFont("Helvetica-Bold", 10)
    tariff_title = "Latest Average Tariff Components"
    if latest_tariff_year:
        tariff_title += f" ({latest_tariff_year})"
    pdf.drawString(24, y, tariff_title)
    y -= 12
    pdf.setFont("Helvetica", 9)
    if latest_tariff_rows:
        for component_name, component_value in latest_tariff_rows[:8]:
            pdf.drawString(28, y, component_name)
            pdf.drawRightString(300, y, f"{component_value:.4f} T$/kWh")
            y -= 11
    else:
        pdf.drawString(28, y, "No tariff data")

    pdf.showPage()
    pdf.save()
    pdf_buffer.seek(0)
    return pdf_buffer.getvalue()


if "summary_pdf_bytes" not in st.session_state:
    st.session_state.summary_pdf_bytes = None

# Row 1: Supply + Offtake (with group titles)
top_supply_col, top_offtake_col = st.columns([2, 3], gap="small")
render_kpi_group(top_supply_col, "Supply", supply_items)
render_kpi_group(top_offtake_col, "Offtake Breakdown", offtake_items)

st.markdown("<div style='height: 0.2rem;'></div>", unsafe_allow_html=True)

# Row 2: Stock by Fuel (left) + Coverage (right)
bottom_left, bottom_right = st.columns([3, 2], gap="small")

with bottom_left:
    if not stock_by_fuel.empty:
        fuel_items = []
        for i, (fuel, stock) in enumerate(stock_by_fuel.items()):
            fuel_items.append((fuel_icon(fuel), str(fuel), f"{format_compact(stock)}", CHART_COLORS[i % len(CHART_COLORS)]))
        render_kpi_group(bottom_left, "Stock by Fuel Type", fuel_items)
    else:
        st.caption("No stock by fuel data")

with bottom_right:
    render_kpi_group(bottom_right, "Coverage", coverage_items)

st.divider()

# Main visualization area
if active_section == "📊 Fuel Supply":
    # Row 1: Key Visualizations
    col1, col2 = st.columns(2)
    
    # Chart 1: Stock Over Time by Fuel Type
    with col1:
        stock_data = filtered_actual.dropna(subset=["Date", "Closing Stock"])
        if not stock_data.empty:
            stock_by_date_fuel = stock_data.groupby(["Date", "Fuel Type"])["Closing Stock"].mean().reset_index()
            fig_stock = px.line(
                stock_by_date_fuel,
                x="Date",
                y="Closing Stock",
                color="Fuel Type",
                title="Stock Level Over Time by Fuel Type",
                markers=True,
                color_discrete_sequence=CHART_COLORS,
            )
            apply_chart_theme(
                fig_stock,
                height=400,
                hovermode="x unified",
                x_title="Date",
                y_title="Closing Stock (L)",
                date_x=True,
            )
            fig_stock.update_layout(title_text="")
            with st.container(border=True, key="chart_stock"):
                render_chart_title(st, "Stock Level Over Time by Fuel Type")
                st.plotly_chart(fig_stock, use_container_width=True)
        else:
            st.info("No stock data available")
    
    # Chart 2: Fuel Offtake Over Time
    with col2:
        offtake_data = filtered_actual.dropna(subset=["Date", "Offtake"])
        if not offtake_data.empty:
            offtake_by_date_fuel = offtake_data.groupby(["Date", "Fuel Type"])["Offtake"].sum().reset_index()
            fig_offtake = px.bar(
                offtake_by_date_fuel,
                x="Date",
                y="Offtake",
                color="Fuel Type",
                title="Daily Offtake (Take-off) by Fuel Type",
                barmode="stack",
                color_discrete_sequence=CHART_COLORS,
            )

            # Add Tonga Power as an explicit overlay series for direct comparison.
            if "Tonga Power Offtake" in offtake_data.columns:
                tonga_power_by_date = (
                    offtake_data.groupby("Date", as_index=False)["Tonga Power Offtake"]
                    .sum(min_count=1)
                    .fillna(0)
                )
                if not tonga_power_by_date.empty and (tonga_power_by_date["Tonga Power Offtake"] > 0).any():
                    fig_offtake.add_trace(
                        go.Scatter(
                            x=tonga_power_by_date["Date"],
                            y=tonga_power_by_date["Tonga Power Offtake"],
                            mode="lines+markers",
                            name="Tonga Power Offtake",
                            line=dict(color="#FDE047", width=2.5),
                            marker=dict(size=6, color="#FDE047"),
                        )
                    )

            apply_chart_theme(
                fig_offtake,
                height=400,
                hovermode="x unified",
                x_title="Date",
                y_title="Offtake (L)",
                date_x=True,
            )
            fig_offtake.update_layout(title_text="")
            with st.container(border=True, key="chart_offtake"):
                render_chart_title(st, "Daily Offtake (Take-off) by Fuel Type")
                st.plotly_chart(fig_offtake, use_container_width=True)
        else:
            st.info("No offtake data available")
    
    st.divider()
    
    # Row 2: Comparative Analysis
    col3, col4 = st.columns(2)
    
    # Chart 3: Stock by Location
    with col3:
        stock_by_location = filtered_actual.dropna(subset=["Location", "Closing Stock"])
        if not stock_by_location.empty:
            latest_stock_loc = stock_by_location.sort_values("Date").drop_duplicates(
                subset=["Location", "Fuel Type"], keep="last"
            )
            stock_loc_summary = latest_stock_loc.groupby(["Location", "Fuel Type"])["Closing Stock"].sum().reset_index()
            fig_loc = px.bar(
                stock_loc_summary,
                x="Location",
                y="Closing Stock",
                color="Fuel Type",
                title="Current Stock by Location",
                barmode="group",
                color_discrete_sequence=CHART_COLORS,
            )
            apply_chart_theme(fig_loc, height=400, x_title="Location", y_title="Closing Stock (L)")
            fig_loc.update_layout(title_text="")
            with st.container(border=True, key="chart_location"):
                render_chart_title(st, "Current Stock by Location")
                st.plotly_chart(fig_loc, use_container_width=True)
        else:
            st.info("No location data available")
    
    # Chart 4: Resupply Schedule
    with col4:
        resupply_data = resupply_df.dropna(subset=["Date", "Quantity"])
        if not resupply_data.empty:
            resupply_summary = resupply_data.groupby(["Date", "Fuel Type"])["Quantity"].sum().reset_index()
            fig_resupply = px.bar(
                resupply_summary,
                x="Date",
                y="Quantity",
                color="Fuel Type",
                title="Scheduled Resupply by Date",
                barmode="stack",
                color_discrete_sequence=CHART_COLORS,
            )
            apply_chart_theme(fig_resupply, height=400, x_title="Resupply Date", y_title="Quantity (L)", date_x=True)
            fig_resupply.update_layout(title_text="")
            with st.container(border=True, key="chart_resupply"):
                render_chart_title(st, "Scheduled Resupply by Date")
                st.plotly_chart(fig_resupply, use_container_width=True)
        else:
            st.info("No resupply data scheduled")

elif active_section == "📦 Terminal Data":
    st.subheader("Terminal Capacities")
    
    # Terminal data filters
    t_locations = sorted([x for x in terminal_df["Location"].dropna().unique()])
    t_info = sorted([x for x in terminal_df["Terminal Info"].dropna().unique()])

    t_col1, t_col2 = st.columns(2)
    t_loc_box = t_col1.container(border=True, key="terminal_location_filter_box")
    t_info_box = t_col2.container(border=True, key="terminal_type_filter_box")
    t_loc_sel = checkbox_slicer_horizontal(t_loc_box, "Terminal Location", t_locations, "t_loc")
    t_info_sel = checkbox_slicer_horizontal(t_info_box, "Terminal Type", t_info, "t_info")
    
    filtered_terminal = terminal_df[
        terminal_df["Company"].isin(company_sel)
        & terminal_df["Location"].isin(t_loc_sel)
        & terminal_df["Terminal Info"].isin(t_info_sel)
    ].copy()
    
    # Terminal visualization
    valid_terminal = filtered_terminal.dropna(subset=["Quantity"])
    if not valid_terminal.empty:
        fig_terminal = px.bar(
            valid_terminal,
            x="Terminal Info",
            y="Quantity",
            color="Fuel Type",
            facet_col="Location",
            title="Terminal Capacities by Location and Type",
            barmode="group",
            color_discrete_sequence=CHART_COLORS,
        )
        apply_chart_theme(fig_terminal, height=450, x_title="Terminal Category", y_title="Capacity (L)")
        fig_terminal.update_layout(title_text="")
        fig_terminal.for_each_annotation(lambda a: a.update(text=a.text.split("=")[-1]))
        with st.container(border=True, key="chart_terminal"):
            render_chart_title(st, "Terminal Capacities by Location and Type")
            st.plotly_chart(fig_terminal, use_container_width=True)
    else:
        st.info("No terminal data available")

    # Company-level comparison of current stock against available terminal capacity.
    stock_company_base = filtered_actual.dropna(subset=["Company", "Date", "Closing Stock"])
    terminal_company_base = filtered_terminal.dropna(subset=["Company", "Quantity"])

    stock_by_company = pd.DataFrame(columns=["Company", "Current Stock (L)"])
    if not stock_company_base.empty:
        latest_stock_company = stock_company_base.sort_values("Date").drop_duplicates(
            subset=["Company", "Location", "Fuel Type"], keep="last"
        )
        stock_by_company = (
            latest_stock_company.groupby("Company", as_index=False)["Closing Stock"]
            .sum()
            .rename(columns={"Closing Stock": "Current Stock (L)"})
        )

    capacity_by_company = pd.DataFrame(columns=["Company", "Terminal Capacity (L)"])
    if not terminal_company_base.empty:
        capacity_by_company = (
            terminal_company_base.groupby("Company", as_index=False)["Quantity"]
            .sum()
            .rename(columns={"Quantity": "Terminal Capacity (L)"})
        )

    company_compare = pd.merge(stock_by_company, capacity_by_company, on="Company", how="outer").fillna(0)

    if not company_compare.empty:
        company_compare["Utilization (%)"] = np.where(
            company_compare["Terminal Capacity (L)"] > 0,
            (company_compare["Current Stock (L)"] / company_compare["Terminal Capacity (L)"]) * 100,
            np.nan,
        )
        company_compare = company_compare.sort_values("Terminal Capacity (L)", ascending=False)

        # Battery-style view: show how full each company's storage is.
        battery_df = company_compare[["Company", "Utilization (%)"]].copy()
        battery_df["Utilization (%)"] = battery_df["Utilization (%)"].clip(lower=0, upper=100)
        battery_df["Remaining (%)"] = 100 - battery_df["Utilization (%)"]
        battery_df["Status"] = np.select(
            [
                battery_df["Utilization (%)"] < 30,
                battery_df["Utilization (%)"].between(30, 60, inclusive="left"),
            ],
            ["Low", "Medium"],
            default="High",
        )

        fig_battery = go.Figure()
        fig_battery.add_trace(
            go.Bar(
                y=battery_df["Company"],
                x=battery_df["Utilization (%)"],
                orientation="h",
                name="Filled",
                marker=dict(
                    color=np.where(
                        battery_df["Status"] == "Low",
                        "#EF4444",
                        np.where(battery_df["Status"] == "Medium", "#F59E0B", "#22C55E"),
                    ),
                ),
                text=[f"{v:.1f}%" for v in battery_df["Utilization (%)"]],
                textposition="inside",
                insidetextanchor="middle",
                hovertemplate="Company=%{y}<br>Fill=%{x:.1f}%<extra></extra>",
            )
        )
        fig_battery.add_trace(
            go.Bar(
                y=battery_df["Company"],
                x=battery_df["Remaining (%)"],
                orientation="h",
                name="Remaining",
                marker=dict(color="rgba(148, 163, 184, 0.35)"),
                hovertemplate="Company=%{y}<br>Remaining=%{x:.1f}%<extra></extra>",
            )
        )
        apply_chart_theme(fig_battery, height=350, x_title="Terminal Fill Level (%)", y_title="Company")
        fig_battery.update_layout(
            barmode="stack",
            title_text="",
            xaxis=dict(range=[0, 100], ticksuffix="%"),
            legend_title_text="",
        )

        with st.container(border=True, key="chart_battery_stock_capacity"):
            render_chart_title(st, "Terminal Fill Level by Company")
            st.plotly_chart(fig_battery, use_container_width=True)
    else:
        st.info("No company data available for stock vs capacity comparison")

    # Location-level comparison and battery/discharge view.
    stock_location_base = filtered_actual.dropna(subset=["Location", "Date", "Closing Stock"])
    terminal_location_base = filtered_terminal.dropna(subset=["Location", "Quantity"])

    stock_by_location = pd.DataFrame(columns=["Location", "Current Stock (L)"])
    if not stock_location_base.empty:
        latest_stock_location = stock_location_base.sort_values("Date").drop_duplicates(
            subset=["Company", "Location", "Fuel Type"], keep="last"
        )
        stock_by_location = (
            latest_stock_location.groupby("Location", as_index=False)["Closing Stock"]
            .sum()
            .rename(columns={"Closing Stock": "Current Stock (L)"})
        )

    capacity_by_location = pd.DataFrame(columns=["Location", "Terminal Capacity (L)"])
    if not terminal_location_base.empty:
        capacity_by_location = (
            terminal_location_base.groupby("Location", as_index=False)["Quantity"]
            .sum()
            .rename(columns={"Quantity": "Terminal Capacity (L)"})
        )

    location_compare = pd.merge(stock_by_location, capacity_by_location, on="Location", how="outer").fillna(0)

    if not location_compare.empty:
        location_compare["Utilization (%)"] = np.where(
            location_compare["Terminal Capacity (L)"] > 0,
            (location_compare["Current Stock (L)"] / location_compare["Terminal Capacity (L)"]) * 100,
            np.nan,
        )
        location_compare = location_compare.sort_values("Terminal Capacity (L)", ascending=False)

        battery_loc_df = location_compare[["Location", "Utilization (%)"]].copy()
        battery_loc_df["Utilization (%)"] = battery_loc_df["Utilization (%)"].clip(lower=0, upper=100)
        battery_loc_df["Remaining (%)"] = 100 - battery_loc_df["Utilization (%)"]
        battery_loc_df["Status"] = np.select(
            [
                battery_loc_df["Utilization (%)"] < 30,
                battery_loc_df["Utilization (%)"].between(30, 60, inclusive="left"),
            ],
            ["Low", "Medium"],
            default="High",
        )

        fig_battery_location = go.Figure()
        fig_battery_location.add_trace(
            go.Bar(
                y=battery_loc_df["Location"],
                x=battery_loc_df["Utilization (%)"],
                orientation="h",
                name="Filled",
                marker=dict(
                    color=np.where(
                        battery_loc_df["Status"] == "Low",
                        "#EF4444",
                        np.where(battery_loc_df["Status"] == "Medium", "#F59E0B", "#22C55E"),
                    ),
                ),
                text=[f"{v:.1f}%" for v in battery_loc_df["Utilization (%)"]],
                textposition="inside",
                insidetextanchor="middle",
                hovertemplate="Location=%{y}<br>Fill=%{x:.1f}%<extra></extra>",
            )
        )
        fig_battery_location.add_trace(
            go.Bar(
                y=battery_loc_df["Location"],
                x=battery_loc_df["Remaining (%)"],
                orientation="h",
                name="Remaining",
                marker=dict(color="rgba(148, 163, 184, 0.35)"),
                hovertemplate="Location=%{y}<br>Remaining=%{x:.1f}%<extra></extra>",
            )
        )
        apply_chart_theme(fig_battery_location, height=320, x_title="Terminal Fill Level (%)", y_title="Location")
        fig_battery_location.update_layout(
            barmode="stack",
            title_text="",
            xaxis=dict(range=[0, 100], ticksuffix="%"),
            legend_title_text="",
        )

        with st.container(border=True, key="chart_battery_stock_capacity_location"):
            render_chart_title(st, "Terminal Fill Level by Location")
            st.plotly_chart(fig_battery_location, use_container_width=True)
    else:
        st.info("No location data available for stock vs capacity comparison")

elif active_section == "💰 Prices & Tariffs":
    if price_df is None or tariff_df is None:
        st.warning(
            "Price & tariff data file not found. "
            "Place Transformed_for_Analysis.xlsx in the project folder."
        )
    else:
        # ── Fuel Prices ───────────────────────────────────────────────
        st.subheader("Fuel Prices")

        latest_price_base = price_df.dropna(subset=["Date", "Price"]).sort_values("Date")
        latest_retail_prices = (
            latest_price_base[latest_price_base["Price_Type"] == "Retail"]
            .drop_duplicates(subset=["Fuel"], keep="last")
        )
        latest_wholesale_prices = (
            latest_price_base[latest_price_base["Price_Type"] == "Wholesale"]
            .drop_duplicates(subset=["Fuel"], keep="last")
        )

        wholesale_col, retail_col = st.columns(2)

        with wholesale_col:
            if not latest_wholesale_prices.empty:
                wholesale_kpi_items = [
                    (
                        fuel_icon(row["Fuel"]),
                        row["Fuel"],
                        f"T${row['Price']:.2f}",
                        CHART_COLORS[i % len(CHART_COLORS)],
                    )
                    for i, (_, row) in enumerate(latest_wholesale_prices.iterrows())
                ]
                render_kpi_group(st, "Latest Wholesale Prices (T$/L)", wholesale_kpi_items)
            else:
                st.info("No wholesale price data available")

        with retail_col:
            if not latest_retail_prices.empty:
                retail_kpi_items = [
                    (
                        fuel_icon(row["Fuel"]),
                        row["Fuel"],
                        f"T${row['Price']:.2f}",
                        CHART_COLORS[i % len(CHART_COLORS)],
                    )
                    for i, (_, row) in enumerate(latest_retail_prices.iterrows())
                ]
                render_kpi_group(st, "Latest Retail Prices (T$/L)", retail_kpi_items)
            else:
                st.info("No retail price data available")

        st.markdown("<div style='height: 0.3rem;'></div>", unsafe_allow_html=True)

        fp_filtered = price_df.copy()
        if price_type_sel:
            fp_filtered = fp_filtered[fp_filtered["Price_Type"].isin(price_type_sel)]
        if fuel_sel_fp:
            fp_filtered = fp_filtered[fp_filtered["Fuel"].isin(fuel_sel_fp)]
        fp_filtered = fp_filtered.dropna(subset=["Date", "Price"]).sort_values("Date")

        if not fp_filtered.empty:
            fig_fp = px.line(
                fp_filtered,
                x="Date",
                y="Price",
                color="Fuel",
                line_dash="Price_Type" if len(price_type_sel) > 1 else None,
                color_discrete_sequence=CHART_COLORS,
            )
            apply_chart_theme(
                fig_fp,
                height=400,
                hovermode="x unified",
                x_title="Date",
                y_title="Price (T$/L)",
                date_x=True,
            )
            fig_fp.update_layout(title_text="")
            with st.container(border=True, key="chart_fuel_price"):
                render_chart_title(st, "Fuel Price Trends Over Time")
                st.plotly_chart(fig_fp, use_container_width=True)
        else:
            st.info("No fuel price data matches the selected filters")

        st.divider()

        # ── Electricity Tariffs ───────────────────────────────────────
        st.subheader("Electricity Tariffs")

        _MONTH_NUM = {
            "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
            "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
        }

        def _parse_tariff_date(row):
            try:
                yr_start = int(str(row["Year"]).split("/")[0])
                m = _MONTH_NUM.get(str(row["Month"]), 1)
                yr = yr_start if m >= 7 else yr_start + 1
                return pd.Timestamp(year=yr, month=m, day=1)
            except Exception:
                return pd.NaT

        tr_work = tariff_df.copy()
        tr_work["Period"] = tr_work.apply(_parse_tariff_date, axis=1)

        main_components = ["Fuel Component", "Non Fuel component", "Total Tariff"]
        tr_main = tr_work[tr_work["Component"].isin(main_components)].dropna(
            subset=["Period", "Value"]
        )

        latest_year = tr_work["Year"].dropna().iloc[-1] if not tr_work.empty else ""
        tr_latest_yr = tr_main[tr_main["Year"] == latest_year]
        if not tr_latest_yr.empty:
            current_tariff_col, average_tariff_col = st.columns(2)

            latest_period = tr_latest_yr["Period"].max()
            tr_current = tr_latest_yr[tr_latest_yr["Period"] == latest_period]
            current_vals = tr_current.groupby("Component")["Value"].mean()
            current_kpi_items = [
                (
                    "⛽" if c == "Fuel Component" else ("🔌" if "non fuel" in c.lower() else "📊"),
                    c,
                    f"T${v:.4f}/kWh",
                    CHART_COLORS[i % len(CHART_COLORS)],
                )
                for i, (c, v) in enumerate(current_vals.items())
            ]
            render_kpi_group(
                current_tariff_col,
                f"Current Tariff — {pd.Timestamp(latest_period).strftime('%b %Y')} (T$/kWh)",
                current_kpi_items,
            )

            avg_vals = tr_latest_yr.groupby("Component")["Value"].mean()
            average_kpi_items = [
                (
                    "⛽" if c == "Fuel Component" else ("🔌" if "non fuel" in c.lower() else "📊"),
                    c,
                    f"T${v:.4f}/kWh",
                    CHART_COLORS[i % len(CHART_COLORS)],
                )
                for i, (c, v) in enumerate(avg_vals.items())
            ]
            render_kpi_group(average_tariff_col, f"Average Tariff — {latest_year} (T$/kWh)", average_kpi_items)
            st.markdown("<div style='height: 0.3rem;'></div>", unsafe_allow_html=True)

        tr_filtered = (
            tr_main[tr_main["Component"].isin(comp_sel)].sort_values("Period")
            if comp_sel
            else tr_main.sort_values("Period")
        )

        if not tr_filtered.empty:
            fig_tr = px.line(
                tr_filtered,
                x="Period",
                y="Value",
                color="Component",
                color_discrete_sequence=CHART_COLORS,
            )
            apply_chart_theme(
                fig_tr,
                height=400,
                hovermode="x unified",
                x_title="Period",
                y_title="Tariff (T$/kWh)",
                date_x=True,
            )
            fig_tr.update_layout(title_text="")
            with st.container(border=True, key="chart_tariff"):
                render_chart_title(st, "Electricity Tariff Components Over Time")
                st.plotly_chart(fig_tr, use_container_width=True)
        else:
            st.info("No tariff data matches the selected filters")

        fp_heatmap = price_df.copy()
        if price_type_sel:
            fp_heatmap = fp_heatmap[fp_heatmap["Price_Type"].isin(price_type_sel)]
        if fuel_sel_fp:
            fp_heatmap = fp_heatmap[fp_heatmap["Fuel"].isin(fuel_sel_fp)]
        fp_heatmap = fp_heatmap.dropna(subset=["Date", "Price"])
        fp_heatmap["Period"] = fp_heatmap["Date"].dt.to_period("M").dt.to_timestamp()
        fp_heatmap["Series"] = fp_heatmap["Price_Type"] + " " + fp_heatmap["Fuel"]
        fp_matrix = fp_heatmap.pivot_table(
            index="Period",
            columns="Series",
            values="Price",
            aggfunc="mean",
        )

        tr_heatmap = tr_main.copy()
        if comp_sel:
            tr_heatmap = tr_heatmap[tr_heatmap["Component"].isin(comp_sel)]
        tr_matrix = tr_heatmap.pivot_table(
            index="Period",
            columns="Component",
            values="Value",
            aggfunc="mean",
        )

        dependency_frame = fp_matrix.join(tr_matrix, how="inner")
        dependency_corr = dependency_frame.corr(numeric_only=True)

        if not fp_matrix.empty and not tr_matrix.empty and not dependency_corr.empty and len(dependency_frame) >= 3:
            fuel_series = fp_matrix.columns.tolist()
            tariff_series = tr_matrix.columns.tolist()
            dependency_view = dependency_corr.loc[fuel_series, tariff_series]

            fig_dependency = go.Figure(
                data=go.Heatmap(
                    z=dependency_view.values,
                    x=dependency_view.columns.tolist(),
                    y=dependency_view.index.tolist(),
                    colorscale="RdBu",
                    zmin=-1,
                    zmax=1,
                    zmid=0,
                    text=np.round(dependency_view.values, 2),
                    texttemplate="%{text}",
                    textfont={"size": 12},
                    colorbar=dict(title="Correlation"),
                    hovertemplate="Fuel Price: %{y}<br>Tariff: %{x}<br>Correlation: %{z:.2f}<extra></extra>",
                )
            )
            fig_dependency.update_layout(
                height=460,
                paper_bgcolor="rgba(0, 0, 0, 0)",
                plot_bgcolor="rgba(12, 18, 26, 0.86)",
                font=dict(color="#E6EDF5", size=13),
                margin=dict(l=16, r=16, t=24, b=16),
                xaxis=dict(title="Tariff Component", side="bottom"),
                yaxis=dict(title="Fuel Price Series"),
            )
            with st.container(border=True, key="chart_dependency"):
                render_chart_title(st, "Fuel Price vs Tariff Dependence Heatmap")
                st.caption("Correlation across overlapping monthly periods. Values closer to 1 move together more strongly; values near -1 move in opposite directions.")
                st.plotly_chart(fig_dependency, use_container_width=True)
        else:
            st.info("Not enough overlapping monthly data to calculate a dependency heatmap")

elif active_section == "🔮 Forecast":
    st.subheader("Fuel Forecast")
    st.caption("Forecast delivery and projected closing stock from Fuel Company reports and manual entries.")

    if filtered_forecast.empty:
        st.info("No forecast data available for the selected filters")
    else:
        upcoming_30 = filtered_forecast[
            (filtered_forecast["Date"] >= pd.Timestamp.now().normalize())
            & (filtered_forecast["Date"] <= pd.Timestamp.now().normalize() + pd.Timedelta(days=30))
        ]
        total_forecast_delivery = upcoming_30["Forecast Delivery"].fillna(0).sum()

        latest_forecast = (
            filtered_forecast.dropna(subset=["Date", "Forecast Closing"])
            .sort_values("Date")
            .drop_duplicates(subset=["Company", "Location", "Fuel Type"], keep="last")
        )
        projected_closing_total = latest_forecast["Forecast Closing"].fillna(0).sum()

        kpi_col1, kpi_col2 = st.columns(2)
        render_kpi_group(
            kpi_col1,
            "Forecast Delivery (Next 30 Days)",
            [("📦", "Total Delivery", f"{format_compact(total_forecast_delivery)}", "#34D399")],
        )
        render_kpi_group(
            kpi_col2,
            "Latest Projected Closing",
            [("🔮", "Projected Stock", f"{format_compact(projected_closing_total)}", "#60A5FA")],
        )

        st.markdown("<div style='height: 0.3rem;'></div>", unsafe_allow_html=True)

        chart_col1, chart_col2 = st.columns(2)

        with chart_col1:
            delivery_series = filtered_forecast.dropna(subset=["Date", "Forecast Delivery"])
            if not delivery_series.empty:
                forecast_delivery_by_date = (
                    delivery_series.groupby(["Date", "Fuel Type"], as_index=False)["Forecast Delivery"].sum()
                )
                fig_forecast_delivery = px.bar(
                    forecast_delivery_by_date,
                    x="Date",
                    y="Forecast Delivery",
                    color="Fuel Type",
                    barmode="stack",
                    color_discrete_sequence=CHART_COLORS,
                )
                apply_chart_theme(
                    fig_forecast_delivery,
                    height=400,
                    x_title="Date",
                    y_title="Forecast Delivery (L)",
                    date_x=True,
                )
                fig_forecast_delivery.update_layout(title_text="")
                with st.container(border=True, key="chart_forecast_delivery"):
                    render_chart_title(st, "Forecast Delivery by Date")
                    st.plotly_chart(fig_forecast_delivery, use_container_width=True)
            else:
                st.info("No forecast delivery values available")

        with chart_col2:
            closing_series = filtered_forecast.dropna(subset=["Date", "Forecast Closing"])
            if not closing_series.empty:
                forecast_closing_by_date = (
                    closing_series.groupby(["Date", "Fuel Type"], as_index=False)["Forecast Closing"].mean()
                )
                fig_forecast_closing = px.line(
                    forecast_closing_by_date,
                    x="Date",
                    y="Forecast Closing",
                    color="Fuel Type",
                    markers=True,
                    color_discrete_sequence=CHART_COLORS,
                )
                apply_chart_theme(
                    fig_forecast_closing,
                    height=400,
                    hovermode="x unified",
                    x_title="Date",
                    y_title="Projected Closing Stock (L)",
                    date_x=True,
                )
                fig_forecast_closing.update_layout(title_text="")
                with st.container(border=True, key="chart_forecast_closing"):
                    render_chart_title(st, "Projected Closing Stock Trend")
                    st.plotly_chart(fig_forecast_closing, use_container_width=True)
            else:
                st.info("No projected closing values available")

        st.divider()
        st.markdown("### Forecast Detail")
        forecast_detail = filtered_forecast.sort_values("Date", ascending=False)[
            ["Date", "Company", "Location", "Fuel Type", "Forecast Delivery", "Forecast Closing", "Source"]
        ]
        st.dataframe(forecast_detail.head(50), use_container_width=True, hide_index=True)

elif active_section == "✏️ Data Entry":
    # Check if user is authenticated and session is valid
    if not check_session_timeout():
        show_login_form()
    else:
        st.subheader("Add New Fuel Data")
        st.caption(f"Logged in as: **{st.session_state.username}** | [Logout](#logout)")
        
        # Initialize session state for forms
        if "entry_mode" not in st.session_state:
            st.session_state.entry_mode = "Actual Stock"
        
        # Tab selection for entry type
        entry_tabs = st.tabs(["📥 Actual Stock", "📦 Resupply Schedule", "💵 Fuel Prices", "🔮 Forecast", "🎯 Tariffs"])

        # Persistent fuel type selection for all forms
        all_fuels = sorted(actual_df["Fuel Type"].dropna().unique())
        if "selected_fuel_type" not in st.session_state:
            st.session_state.selected_fuel_type = all_fuels[0] if all_fuels else ""
        
        with entry_tabs[0]:
            st.markdown("### Record Daily Fuel Stock")
            st.caption("Add a new daily stock entry for a fuel company location.")
            # Show most recent entry from this session (handle both single and multiple rows)
            if "last_actual_entry" in st.session_state:
                last_entry = st.session_state["last_actual_entry"]
                if isinstance(last_entry, list):
                    df_last = pd.DataFrame(last_entry)
                else:
                    df_last = pd.DataFrame([last_entry])
                st.info("Most recent entry:")
                st.dataframe(df_last, use_container_width=True, hide_index=True)

            with st.form("actual_stock_form", border=True):
                left, right = st.columns([1.2, 2.2])
                with left:
                    entry_date = st.date_input("Stock Date", key="actual_date")
                    entry_company = st.selectbox(
                        "Company",
                        options=sorted(actual_df["Company"].dropna().unique()),
                        key="actual_company"
                    )
                    entry_location = st.selectbox(
                        "Location",
                        options=sorted(actual_df["Location"].dropna().unique()),
                        key="actual_location"
                    )
                    update_existing_actual = st.checkbox(
                        "Update existing row if same Date + Company + Location + Fuel Type exists",
                        value=True,
                        key="actual_update_existing",
                    )
                with right:
                    st.markdown("#### Enter Stock and Offtake for Each Fuel Type")
                    with st.container(border=True):
                        st.markdown("<style>div[data-testid='column'] label {font-weight: 500;}</style>", unsafe_allow_html=True)
                        # Add header row for labels
                        header_cols = st.columns([1.2, 1.2, 1.2])
                        header_cols[0].markdown("<span style='font-weight:600'></span>", unsafe_allow_html=True)
                        header_cols[1].markdown("<span style='font-weight:600'>Closing Stock (L)</span>", unsafe_allow_html=True)
                        header_cols[2].markdown("<span style='font-weight:600'>Offtake (L)</span>", unsafe_allow_html=True)
                        fuel_rows = [
                            ("Petrol", "actual_petrol_closing", "actual_petrol_offtake", "actual_petrol_tonga"),
                            ("Diesel", "actual_diesel_closing", "actual_diesel_offtake", "actual_diesel_tonga"),
                            ("Kerosene", "actual_kerosene_closing", "actual_kerosene_offtake", "actual_kerosene_tonga"),
                        ]
                        fuel_entries = []
                        for fuel, closing_key, offtake_key, tonga_key in fuel_rows:
                            cols = st.columns([1.2, 1.2, 1.2])
                            cols[0].markdown(f"<span style='font-weight:600'>{fuel}</span>", unsafe_allow_html=True)
                            closing = cols[1].number_input(f"Closing Stock (L) [{fuel}]", min_value=0, value=0, key=closing_key, label_visibility="collapsed")
                            offtake = cols[2].number_input(f"Offtake (L) [{fuel}]", min_value=0, value=0, key=offtake_key, label_visibility="collapsed")
                            tonga = None
                            # Only show Tonga Power Offtake for TotalEnergies
                            if fuel == "Diesel" and entry_company.strip().lower().replace(" ","") in ["totalenergies","totalenergiesmarketing"]:
                                tonga = st.number_input(f"Tonga Power Offtake (L) [{fuel}]", min_value=0, value=0, key=tonga_key)
                            fuel_entries.append({
                                "Fuel Type": fuel,
                                "Closing Stock": closing,
                                "Offtake": offtake,
                                "Tonga Power Offtake": tonga if tonga is not None else 0,
                            })
                    submitted = st.form_submit_button("➕ Add Stock Entry", use_container_width=True)

                if submitted:
                    # Save each fuel row as a separate entry
                    last_entries = []
                    for entry in fuel_entries:
                        # Only save if at least one value is entered
                        if entry["Closing Stock"] > 0 or entry["Offtake"] > 0 or (entry["Tonga Power Offtake"] and entry["Tonga Power Offtake"] > 0):
                            success, message = save_actual_data(
                                str(file_to_use),
                                entry_date,
                                entry_company,
                                entry_location,
                                entry["Fuel Type"],
                                entry["Closing Stock"],
                                entry["Offtake"],
                                entry["Tonga Power Offtake"],
                                allow_update=update_existing_actual,
                            )
                            last_entries.append({
                                "Date": entry_date,
                                "Company": entry_company,
                                "Location": entry_location,
                                "Fuel Type": entry["Fuel Type"],
                                "Closing Stock": entry["Closing Stock"],
                                "Offtake": entry["Offtake"],
                                "Tonga Power Offtake": entry["Tonga Power Offtake"],
                                "Status": "Success" if success else "Failed",
                                "Message": message,
                            })
                    if last_entries:
                        st.session_state["last_actual_entry"] = last_entries
                        st.success("Entries saved. See most recent above.")
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.warning("No values entered for any fuel type.")
        
        with entry_tabs[1]:
            st.markdown("### Schedule Fuel Resupply")
            st.caption("Add a new scheduled resupply event for a fuel company location.")
            if "last_resupply_entry" in st.session_state:
                st.info("Most recent entry:")
                st.dataframe(pd.DataFrame([st.session_state["last_resupply_entry"]]), use_container_width=True, hide_index=True)
            
            with st.form("resupply_form", border=True):
                left, right = st.columns([1.2, 2.2])
                with left:
                    resupply_date = st.date_input("Resupply Date", key="resupply_date")
                    resupply_company = st.selectbox(
                        "Company",
                        options=sorted(actual_df["Company"].dropna().unique()),
                        key="resupply_company"
                    )
                    resupply_location = st.selectbox(
                        "Location",
                        options=sorted(actual_df["Location"].dropna().unique()),
                        key="resupply_location"
                    )
                    update_existing_resupply = st.checkbox(
                        "Update existing row if same Date + Company + Location + Fuel Type exists",
                        value=True,
                        key="resupply_update_existing",
                    )
                with right:
                    st.markdown("#### Enter Resupply Quantity for Each Fuel Type")
                    with st.container(border=True):
                        st.markdown("<style>div[data-testid='column'] label {font-weight: 500;}</style>", unsafe_allow_html=True)
                        resupply_fuel_rows = [
                            ("Petrol", "resupply_petrol_quantity"),
                            ("Diesel", "resupply_diesel_quantity"),
                            ("Kerosene", "resupply_kerosene_quantity"),
                        ]
                        resupply_entries = []
                        for fuel, quantity_key in resupply_fuel_rows:
                            cols = st.columns([1.2, 2])
                            cols[0].markdown(f"<span style='font-weight:600'>{fuel}</span>", unsafe_allow_html=True)
                            quantity = cols[1].number_input(f"Quantity (L) [{fuel}]", min_value=0, value=0, key=quantity_key, label_visibility="collapsed")
                            resupply_entries.append({
                                "Fuel Type": fuel,
                                "Quantity": quantity,
                            })
                    submitted_resupply = st.form_submit_button("➕ Add Resupply Entry", use_container_width=True)

                if submitted_resupply:
                    last_resupply_entries = []
                    for entry in resupply_entries:
                        if entry["Quantity"] > 0:
                            success, message = save_resupply_data(
                                str(file_to_use),
                                resupply_date,
                                resupply_company,
                                resupply_location,
                                entry["Fuel Type"],
                                entry["Quantity"],
                                allow_update=update_existing_resupply,
                            )
                            last_resupply_entries.append({
                                "Date": resupply_date,
                                "Company": resupply_company,
                                "Location": resupply_location,
                                "Fuel Type": entry["Fuel Type"],
                                "Quantity": entry["Quantity"],
                            })
                    if last_resupply_entries:
                        st.session_state["last_resupply_entry"] = last_resupply_entries
                        st.success("Entries saved. See most recent above.")
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.warning("No values entered for any fuel type.")
                
                if submitted_resupply:
                    success, message = save_resupply_data(
                        str(file_to_use),
                        resupply_date,
                        resupply_company,
                        resupply_location,
                        resupply_fuel,
                        resupply_quantity,
                        allow_update=update_existing_resupply,
                    )
                    st.session_state["last_resupply_entry"] = {
                        "Date": resupply_date,
                        "Company": resupply_company,
                        "Location": resupply_location,
                        "Fuel Type": resupply_fuel,
                        "Quantity": resupply_quantity,
                    }
                    if success:
                        st.success(message)
                        st.info("💡 The dashboard will automatically reload with the new data on your next view refresh.")
                    else:
                        st.error(message)
        
        with entry_tabs[2]:
            st.markdown("### Record Fuel Price")
            st.caption("Add a new fuel price entry for tracking price trends.")
            
            with st.form("fuel_price_form", border=True):
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    price_date = st.date_input("Price Date", key="price_date")
                
                with col2:
                    price_fuel = st.selectbox(
                        "Fuel Type",
                        options=["Petrol", "Diesel", "Heavy Fuel Oil"],
                        key="price_fuel"
                    )
                
                with col3:
                    price_type_sel = st.selectbox(
                        "Price Type",
                        options=["Wholesale", "Retail", "Import"],
                        key="price_type_sel"
                    )
                
                col4, col5 = st.columns(2)
                
                with col4:
                    price_value = st.number_input(
                        "Price (per Liter)",
                        min_value=0.0,
                        value=0.0,
                        step=0.01,
                        key="price_value"
                    )
                
                with col5:
                    price_currency = st.selectbox(
                        "Currency",
                        options=["TOP", "USD", "AUD"],
                        key="price_currency"
                    )
                
                submitted_price = st.form_submit_button("➕ Add Fuel Price", use_container_width=True)
                
                if submitted_price:
                    if PRICE_FILE.exists():
                        success, message = save_fuel_price(
                            str(PRICE_FILE),
                            price_date,
                            price_fuel,
                            price_type_sel,
                            price_value
                        )
                        if success:
                            st.success(message)
                            st.info("💡 The dashboard will automatically reload with the new data on your next view refresh.")
                        else:
                            st.error(message)
                    else:
                        st.error("Price data file not found. Please check the system configuration.")
        
        with entry_tabs[3]:
            st.markdown("### Record Forecast")
            st.caption("Add a manual forecast delivery/closing entry.")
            if "last_forecast_entry" in st.session_state:
                st.info("Most recent entry:")
                st.dataframe(pd.DataFrame([st.session_state["last_forecast_entry"]]), use_container_width=True, hide_index=True)

            with st.form("forecast_form", border=True):
                left, right = st.columns([1.2, 2.2])
                with left:
                    forecast_date = st.date_input("Forecast Date", key="forecast_date")
                    forecast_company = st.selectbox(
                        "Company",
                        options=companies if companies else sorted(actual_df["Company"].dropna().unique()),
                        key="forecast_company"
                    )
                    forecast_location = st.selectbox(
                        "Location",
                        options=locations if locations else sorted(actual_df["Location"].dropna().unique()),
                        key="forecast_location"
                    )
                    update_existing_forecast = st.checkbox(
                        "Update existing row if same Date + Company + Location + Fuel Type exists",
                        value=True,
                        key="forecast_update_existing",
                    )
                with right:
                    st.markdown("#### Enter Forecast Delivery and Closing for Each Fuel Type")
                    with st.container(border=True):
                        st.markdown("<style>div[data-testid='column'] label {font-weight: 500;}</style>", unsafe_allow_html=True)
                        forecast_fuel_rows = [
                            ("Petrol", "forecast_petrol_delivery", "forecast_petrol_closing"),
                            ("Diesel", "forecast_diesel_delivery", "forecast_diesel_closing"),
                            ("Kerosene", "forecast_kerosene_delivery", "forecast_kerosene_closing"),
                        ]
                        forecast_entries = []
                        for fuel, delivery_key, closing_key in forecast_fuel_rows:
                            cols = st.columns([1.2, 1.2, 1.2])
                            cols[0].markdown(f"<span style='font-weight:600'>{fuel}</span>", unsafe_allow_html=True)
                            delivery = cols[1].number_input(f"Forecast Delivery (L) [{fuel}]", min_value=0.0, value=0.0, step=1.0, key=delivery_key, label_visibility="collapsed")
                            closing = cols[2].number_input(f"Forecast Closing (L) [{fuel}]", min_value=0.0, value=0.0, step=1.0, key=closing_key, label_visibility="collapsed")
                            forecast_entries.append({
                                "Fuel Type": fuel,
                                "Forecast Delivery": delivery,
                                "Forecast Closing": closing,
                            })
                    submitted_forecast = st.form_submit_button("➕ Add Forecast Entry", use_container_width=True)

                if submitted_forecast:
                    last_forecast_entries = []
                    for entry in forecast_entries:
                        if entry["Forecast Delivery"] > 0 or entry["Forecast Closing"] > 0:
                            success, message = save_forecast_entry(
                                str(FORECAST_ENTRY_FILE),
                                forecast_date,
                                forecast_company,
                                forecast_location,
                                entry["Fuel Type"],
                                entry["Forecast Delivery"],
                                entry["Forecast Closing"],
                                allow_update=update_existing_forecast,
                            )
                            last_forecast_entries.append({
                                "Date": forecast_date,
                                "Company": forecast_company,
                                "Location": forecast_location,
                                "Fuel Type": entry["Fuel Type"],
                                "Forecast Delivery": entry["Forecast Delivery"],
                                "Forecast Closing": entry["Forecast Closing"],
                            })
                    if last_forecast_entries:
                        st.session_state["last_forecast_entry"] = last_forecast_entries
                        st.success("Entries saved. See most recent above.")
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.warning("No values entered for any fuel type.")

        with entry_tabs[4]:
            st.markdown("### Record Tariff")
            st.caption("Add a new tariff component entry for tracking tariff changes.")
            
            with st.form("tariff_form", border=True):
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    tariff_date = st.date_input("Tariff Date", key="tariff_date")
                
                with col2:
                    tariff_component = st.selectbox(
                        "Tariff Component",
                        options=["Base Rate", "Excise Tax", "Distribution Cost", "Administration Fee", "Environmental Levy"],
                        key="tariff_component"
                    )
                
                with col3:
                    tariff_period = st.selectbox(
                        "Period",
                        options=["Monthly", "Quarterly", "Annual"],
                        key="tariff_period"
                    )
                
                col4, col5 = st.columns(2)
                
                with col4:
                    tariff_value = st.number_input(
                        "Tariff Value",
                        min_value=0.0,
                        value=0.0,
                        step=0.01,
                        key="tariff_value"
                    )
                
                with col5:
                    tariff_unit = st.selectbox(
                        "Unit",
                        options=["TOP/L", "TOP/Unit", "Percentage"],
                        key="tariff_unit"
                    )
                
                submitted_tariff = st.form_submit_button("➕ Add Tariff Entry", use_container_width=True)
                
                if submitted_tariff:
                    if PRICE_FILE.exists():
                        success, message = save_tariff(
                            str(PRICE_FILE),
                            tariff_date,
                            tariff_component,
                            tariff_period,
                            tariff_value
                        )
                        if success:
                            st.success(message)
                            st.info("💡 The dashboard will automatically reload with the new data on your next view refresh.")
                        else:
                            st.error(message)
                    else:
                        st.error("Price data file not found. Please check the system configuration.")
        
        st.divider()
        st.markdown("### Recent Data Entries")
        st.caption("Latest entries for all data types currently in the system.")
        
        tab_recent1, tab_recent2, tab_recent3, tab_recent4, tab_recent5 = st.tabs(["Recent Actual Stock", "Recent Resupply", "Recent Fuel Prices", "Recent Forecast", "Recent Tariffs"])
        
        with tab_recent1:
            try:
                if not actual_df.empty:
                    display_cols = [c for c in ["Date", "Company", "Location", "Fuel Type", "Closing Stock", "Offtake"] if c in actual_df.columns]
                    recent_actual = actual_df.nlargest(10, "Date")[display_cols]
                    st.dataframe(recent_actual, use_container_width=True, hide_index=True)
                else:
                    st.info("No actual stock data available")
            except Exception as e:
                st.error(f"Error displaying recent actual stock: {e}")
        
        with tab_recent2:
            recent_resupply = resupply_df.nlargest(10, "Date")[["Date", "Company", "Location", "Fuel Type", "Quantity"]]
            st.dataframe(recent_resupply, use_container_width=True, hide_index=True)
        
        with tab_recent3:
            if price_df is not None and not price_df.empty:
                recent_prices = price_df.nlargest(10, "Date")[["Date", "Fuel", "Price_Type", "Price"]]
                st.dataframe(recent_prices, use_container_width=True, hide_index=True)
            else:
                st.info("No fuel price data available")
        
        with tab_recent4:
            if not forecast_df.empty:
                recent_forecast = forecast_df.nlargest(10, "Date")[["Date", "Company", "Location", "Fuel Type", "Forecast Delivery", "Forecast Closing", "Source"]]
                st.dataframe(recent_forecast, use_container_width=True, hide_index=True)
            else:
                st.info("No forecast data available")

        with tab_recent5:
            if tariff_df is not None and not tariff_df.empty:
                tariff_recent = tariff_df.copy()
                if "Date" in tariff_recent.columns:
                    tariff_recent = tariff_recent.sort_values("Date", ascending=False)
                elif "Year" in tariff_recent.columns:
                    tariff_recent = tariff_recent.sort_values("Year", ascending=False)

                tariff_cols = [col for col in ["Date", "Month", "Year", "Component", "Value"] if col in tariff_recent.columns]
                recent_tariffs = tariff_recent.head(10)[tariff_cols]
                st.dataframe(recent_tariffs, use_container_width=True, hide_index=True)
            else:
                st.info("No tariff data available")


st.divider()
pdf_col1, pdf_col2 = st.columns([1, 5])
with pdf_col1:
    if st.button("⬇️ Generate Summary PDF", key="generate_pdf_btn"):
        st.session_state.summary_pdf_bytes = build_summary_pdf_bytes()
    if st.session_state.summary_pdf_bytes:
        st.download_button(
            "⬇️ Download Summary PDF",
            data=st.session_state.summary_pdf_bytes,
            file_name=f"fuel_dashboard_summary_{pd.Timestamp.now().strftime('%Y%m%d')}.pdf",
            mime="application/pdf",
            key="download_summary_pdf",
        )

with pdf_col2:
    if st.session_state.summary_pdf_bytes:
        st.caption("PDF ready — click Download to save.")
    else:
        st.caption("Click Generate to build the summary PDF.")

st.markdown(
    f"""
    <div style="
        margin-top: 0.6rem;
        padding: 0.52rem 0.75rem;
        background: rgba(229, 231, 235, 0.95);
        border: 1px solid rgba(209, 213, 219, 0.95);
        color: #4B5563;
        font-size: 0.88rem;
        line-height: 1.35;
        display: flex;
        justify-content: center;
        align-items: center;
        gap: 0.7rem;
        flex-wrap: wrap;
        text-align: center;
        border-radius: 12px;
    ">
        {footer_logo_html}
        <span>Developed by Department of Energy under Ministry of MEIDECC for Staged Energy and Fuel Supply Plan Monitoring | Data updated: {last_sync}</span>
    </div>
    """,
    unsafe_allow_html=True,
)
