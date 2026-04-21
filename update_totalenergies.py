import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# Data from the provided image for TotalEnergies
# Dates: 2020-04-13 to 2020-04-19 (assuming year 2020 for April)
data = [
    ["2020-04-13", 524021, 38810, 998251, 165830, 0],
    ["2020-04-14", 524021, 0, 877834, 120000, 0],
    ["2020-04-15", 491471, 32200, 836660, 20000, 19500],
    ["2020-04-16", 491471, 0, 734794, 100000, 0],
    ["2020-04-17", 463521, 27200, 533261, 201010, 0],
    ["2020-04-18", 463521, 0, 533261, 0, 0],
    ["2020-04-19", 463521, 0, 533261, 0, 0],
]

columns = ["Date", "Closing Stock", "Offtake", "Opening Stock", "Resupply", "Tonga Power Offtake"]
df = pd.DataFrame(data, columns=columns)
df["Date"] = pd.to_datetime(df["Date"])

# Path to the Excel file and sheet name
excel_path = "Oil_Data_Consolidated.xlsx"
sheet_name = "Actual"

# Load workbook and sheet
wb = load_workbook(excel_path)
ws = wb[sheet_name]

# Find the header row and column indices
header = [cell.value for cell in ws[1]]
col_idx = {col: header.index(col) + 1 for col in columns if col in header}

# Update or append rows for TotalEnergies for the given dates
for _, row in df.iterrows():
    date_str = row["Date"].strftime("%Y-%m-%d")
    found = False
    for ws_row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws_date = ws_row[col_idx["Date"] - 1].value
        ws_company = ws_row[header.index("Company")].value
        if ws_date and ws_company:
            if pd.to_datetime(ws_date).strftime("%Y-%m-%d") == date_str and str(ws_company).strip().lower() in ["totalenergies", "company 1", "co 1", "co1", "total energies"]:
                # Update values
                ws_row[col_idx["Closing Stock"] - 1].value = row["Closing Stock"]
                ws_row[col_idx["Offtake"] - 1].value = row["Offtake"]
                if "Tonga Power Offtake" in col_idx:
                    ws_row[col_idx["Tonga Power Offtake"] - 1].value = row["Tonga Power Offtake"]
                found = True
                break
    if not found:
        # Append new row
        new_row = [None] * len(header)
        new_row[header.index("Date")] = row["Date"]
        new_row[header.index("Company")] = "TotalEnergies"
        new_row[header.index("Closing Stock")] = row["Closing Stock"]
        new_row[header.index("Offtake")] = row["Offtake"]
        if "Tonga Power Offtake" in header:
            new_row[header.index("Tonga Power Offtake")] = row["Tonga Power Offtake"]
        ws.append(new_row)

wb.save(excel_path)
print("TotalEnergies data updated successfully.")
